from typing import List, Dict, Any, TypedDict, Annotated, Optional, Tuple
from dataclasses import dataclass, asdict
from dotenv import load_dotenv
from langgraph.graph.message import add_messages
from langchain_core.messages import BaseMessage, AIMessage
from langgraph.graph import StateGraph, END
import math

import os
import time
import asyncio
import json
from pathlib import Path

import urllib.request as libreq
import xml.etree.ElementTree as ET
import re
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
import sys

#from test_files.test_model_researcher_streaming import test_model_suggestion_streaming

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# Suppress TensorFlow logging at the module level


# Add missing imports
try:
    import faiss
except ImportError:
    print("Warning: FAISS not installed. Semantic search features will be unavailable.")

# ===== Moved from shared_constants.py =====

@dataclass
class Evidence:
    snippet: str
    source: str
    score: float


@dataclass
class PropertyHit:
    name: str
    evidence: List[Evidence]
    
    @property
    def confidence(self) -> float:
        """Calculate confidence based on evidence."""
        if not self.evidence:
            return 0.0
        
        # Calculate base confidence using independent signals
        prod = 1.0
        for ev in self.evidence:
            prod *= (1.0 - max(0.0, min(1.0, ev.score)))
        base_confidence = 1.0 - prod
        
        # Apply evidence count bonus with diminishing returns
        evidence_bonus = min(0.05 * math.log(len(self.evidence) + 1), 0.15)
        
        final_confidence = min(1.0, base_confidence + evidence_bonus)
        return round(final_confidence, 3)
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary format."""
        return {
            "name": self.name,
            "confidence": self.confidence,
            "evidence": [asdict(ev) for ev in self.evidence],
        }


class BaseState(TypedDict):
    """Base state for all workflows."""
    messages: Annotated[List[BaseMessage], add_messages]
    original_prompt: str  # Pure user query without uploaded data
    uploaded_data: List[str]  # Uploaded file contents as separate field
    current_step: str
    errors: List[str]






# ==================================================================================
# STREAMWRITER HELPER FUNCTION
# ==================================================================================
from langgraph.config import get_stream_writer
def _write_stream(message: str, key: str = "status"):
    """Helper function to write to StreamWriter if available."""
    try:
        # Use LangGraph's get_stream_writer() without parameters (proper way)
        writer = get_stream_writer()
        writer({key: message})
    except Exception:
        # Fallback: try to get stream from config (for testing compatibility)
        try:
            # This fallback is for test compatibility only
            import inspect
            frame = inspect.currentframe()
            while frame:
                if 'config' in frame.f_locals and frame.f_locals['config']:
                    config = frame.f_locals['config']
                    stream = config.get("configurable", {}).get("stream")
                    if stream and hasattr(stream, 'write'):
                        stream.write(message)
                        return
                frame = frame.f_back
        except Exception:
            pass
        # Final fallback: silently fail
        pass



# ===== End moved from shared_constants.py =====

def _clean_text_for_utf8(text):
    """Clean text to ensure UTF-8 compatibility by removing surrogate characters."""
    if not isinstance(text, str):
        return str(text)
    
    # Remove surrogate characters that cause UTF-8 encoding issues
    # Remove surrogate pairs (Unicode range U+D800-U+DFFF)
    text = re.sub(r'[\ud800-\udfff]', '', text)
    
    # Replace other problematic Unicode characters with safe alternatives
    text = text.encode('utf-8', errors='ignore').decode('utf-8')
    
    # Clean up any remaining control characters except newlines and tabs
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', text)
    
    return text


def _load_text_file_safely(file_path: str) -> Tuple[List[str], List[str]]:
    """Attempt to load file content as UTF-8 and gracefully skip binary data.

    Returns a tuple of (text_snippets, warnings)."""

    warnings: List[str] = []
    if not file_path:
        return [], warnings

    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {file_path}")

    suffix = path.suffix.lower()

    if suffix in {".xlsx", ".xls", ".xlsm"}:
        try:
            import pandas as pd  # type: ignore
        except ImportError:
            warnings.append(
                "⚠️ Unable to parse Excel file because pandas is not installed; skipping attachment content."
            )
            return [], warnings

        try:
            excel_data = pd.read_excel(path, sheet_name=None)
        except Exception as exc:  # pragma: no cover - pandas-specific errors
            warnings.append(
                f"⚠️ Failed to parse Excel file '{path.name}': {exc}; skipping attachment content."
            )
            return [], warnings

        text_snippets: List[str] = []
        workbook = None
        openpyxl_failed = False

        for sheet_name, sheet_df in excel_data.items():
            usable_df = None
            if sheet_df is not None:
                trimmed_df = sheet_df.dropna(how="all")
                if not trimmed_df.empty:
                    trimmed_df = trimmed_df.loc[:, trimmed_df.notna().any()]
                if not trimmed_df.empty:
                    usable_df = trimmed_df

            if usable_df is None or usable_df.empty:
                try:
                    alt_df = pd.read_excel(path, sheet_name=sheet_name, header=None)
                    alt_df = alt_df.dropna(how="all")
                    if not alt_df.empty:
                        alt_df = alt_df.loc[:, alt_df.notna().any()]
                except Exception:  # pragma: no cover - pandas-specific errors
                    alt_df = None
                if alt_df is not None and not alt_df.empty:
                    usable_df = alt_df
                    warnings.append(
                        f"ℹ️ Excel sheet '{sheet_name}' parsed without headers due to sparse data."
                    )

            if usable_df is None or usable_df.empty:
                if workbook is None and not openpyxl_failed:
                    try:
                        from openpyxl import load_workbook  # type: ignore

                        workbook = load_workbook(path, read_only=True, data_only=True)
                    except Exception:  # pragma: no cover - openpyxl specific errors
                        workbook = None
                        openpyxl_failed = True

                if workbook is not None and sheet_name in workbook.sheetnames:
                    rows: List[str] = []
                    for row in workbook[sheet_name].iter_rows(values_only=True):
                        if not row or all(
                            (cell is None)
                            or (isinstance(cell, str) and not cell.strip())
                            for cell in row
                        ):
                            continue
                        row_values = ["" if cell is None else str(cell) for cell in row]
                        rows.append(",".join(row_values))
                        if len(rows) >= 200:
                            break
                    if rows:
                        raw_text = "\n".join(rows)
                        cleaned_rows = _clean_text_for_utf8(raw_text)
                        snippet = f"Sheet: {sheet_name}\n{cleaned_rows}"
                        if len(snippet) > 10000:
                            snippet = snippet[:10000] + "\n... (truncated)"
                            warnings.append(
                                f"ℹ️ Excel sheet '{sheet_name}' truncated to 10k characters."
                            )
                        warnings.append(
                            f"ℹ️ Excel sheet '{sheet_name}' extracted via openpyxl fallback."
                        )
                        text_snippets.append(snippet)
                        continue

            if usable_df is None or usable_df.empty:
                continue

            preview_df = usable_df.head(200)
            if len(usable_df) > len(preview_df):
                warnings.append(
                    f"ℹ️ Excel sheet '{sheet_name}' large; attached first {len(preview_df)} rows only."
                )
            preview = preview_df.to_csv(index=False)
            cleaned = _clean_text_for_utf8(preview)
            snippet = f"Sheet: {sheet_name}\n{cleaned}"
            if len(snippet) > 10000:
                snippet = snippet[:10000] + "\n... (truncated)"
                warnings.append(
                    f"ℹ️ Excel sheet '{sheet_name}' truncated to 10k characters."
                )
            text_snippets.append(snippet)

        if not text_snippets:
            warnings.append(
                f"ℹ️ Excel file '{path.name}' contained no tabular data to attach."
            )

        return text_snippets, warnings

    if suffix in {".csv", ".tsv"}:
        try:
            import pandas as pd  # type: ignore
        except ImportError:
            warnings.append(
                "⚠️ Unable to parse CSV/TSV file because pandas is not installed; falling back to raw text."
            )
        else:
            sep = "\t" if suffix == ".tsv" else ","
            try:
                df = pd.read_csv(path, sep=sep)
            except UnicodeDecodeError:
                try:
                    df = pd.read_csv(path, sep=sep, encoding="latin-1")
                    warnings.append(
                        "⚠️ CSV decoding failed with UTF-8; loaded using latin-1 encoding."
                    )
                except Exception as exc:
                    warnings.append(
                        f"⚠️ Failed to parse CSV/TSV file '{path.name}': {exc}; falling back to raw text."
                    )
                    df = None
            except Exception as exc:  # pragma: no cover - pandas-specific errors
                warnings.append(
                    f"⚠️ Failed to parse CSV/TSV file '{path.name}': {exc}; falling back to raw text."
                )
                df = None

            if df is not None:
                preview_df = df.head(200)
                if len(df) > len(preview_df):
                    warnings.append(
                        f"ℹ️ CSV/TSV file '{path.name}' large; attached first {len(preview_df)} rows only."
                    )
                preview_text = preview_df.to_csv(index=False)
                cleaned = _clean_text_for_utf8(preview_text)
                
                return [cleaned], warnings

    try:
        with path.open("r", encoding="utf-8") as handle:
            return [_clean_text_for_utf8(handle.read())], warnings
    except UnicodeDecodeError:
        try:
            size = path.stat().st_size
        except OSError:
            size = None
        size_note = f" ({size} bytes)" if size is not None else ""
        warnings.append(
            f"⚠️ Detected non-text input file '{path.name}'{size_note}; skipping attachment content."
        )
    except FileNotFoundError:
        raise FileNotFoundError(f"Input file not found: {file_path}")
    except Exception as exc:
        raise Exception(f"Error reading input file: {exc}")

    return [], warnings


def format_search_string(query):
    """Format search query for arXiv API."""
    formatted = query.replace('/', ' AND ')
    # URL encode the query
    import urllib.parse
    return urllib.parse.quote(formatted)

    
class ExperimentSuggestionState(BaseState):
    """State object for the experiment suggestion workflow."""
    # Input data
    experimental_results: Dict[str, Any]      # Raw experimental data
    findings_analysis: Dict[str, Any]         # Analysis of current findings
    research_context: Dict[str, Any]          # Context about the research domain
    
    # Processing state
    analysis_completed: bool                  # Whether initial analysis is done
    experiment_categories: List[str]          # Types of experiments identified
    experiment_papers: List[Dict[str, Any]]   # Papers retrieved for experimental guidance
    experiment_search_query: str              # Query used for paper search
    experiment_search_iteration: int          # Current search iteration count
    experiment_validation_results: Dict[str, Any]  # Results from experiment validation (not paper validation)
    experiment_paper_validation_decision: str # Decision from validation (continue/search_new/search_backup)
    experiment_validation_decision: str       # Overall validation decision (PASS/FAIL)
    experiment_iterations: List[Dict[str, Any]]  # History of experiment iterations
    research_direction: Dict[str, Any]        # Research direction analysis
    validated_experiment_papers: List[Dict[str, Any]]  # Validated papers for suggestions
    distilled_methodologies: Dict[str, Any]         # Distilled methodology content from papers
    current_experiment_iteration: int        # Current iteration of experiment suggestion
    iteration_from_state: int                 # Iteration number from state
    analysis_iterations: List[Dict[str, Any]] # Track analysis validation iterations (history)
    direction_iterations: List[Dict[str, Any]] # Track research direction validation iterations (history)
    # Issue tracking for iterative improvement
    past_fixed_issues: List[str]              # Issues that were resolved in previous iterations
    past_unresolved_issues: List[str]         # Issues that persist across iterations
    most_recent_generation_issues: List[str]  # Issues from the most recent experiment generation
    cumulative_validation_feedback: List[Dict[str, Any]]  # Historical validation feedback
    
    # 🆕 PAST MISTAKES TRACKING FOR ITERATIVE LEARNING
    past_experiment_mistakes: List[Dict[str, Any]]  # Historical validation failures for LLM learning
    
    # Output
    experiment_suggestions: str                # Comprehensive experiment suggestions
    experiment_summary: Dict[str, Any]         # Summary of experiment generation
    next_node: str                            # Next node to route to in workflow
    literature_context: str                    # Extracted literature context for experiments
    suggestion_source: str                     # Source of the experiment suggestions
    prioritized_experiments: List[Dict[str, Any]]  # Ranked experiment list
    implementation_roadmap: Dict[str, Any]    # Step-by-step implementation plan
    final_outputs: Dict[str, str]             # Final formatted outputs  

    # Dependencies needed by workflow nodes
    client: Any                                # OpenAI client
    model: str                                 # Model name
    arxiv_processor: Any                       # ArxivPaperProcessor instance


# ==================================================================================
# EXPERIMENT SUGGESTION WORKFLOW NODES
# ==================================================================================
# HELPER FUNCTIONS
# ==================================================================================

def _combine_query_and_data(user_query: str, uploaded_data: List[str]) -> str:
    """Combine user query with uploaded data for LLM prompts when needed."""
    if not uploaded_data:
        return user_query
    
    combined = user_query
    combined += "\n\nAttached Contexts:\n" + "\n\n".join(uploaded_data)
    return combined

def _get_query_for_ranking(user_query: str, uploaded_data: List[str]) -> str:
    """Get the appropriate query for paper ranking - uses pure user query for better relevance."""
    return user_query  # For ranking, we typically want just the user query without uploaded data

def create_experiment_ranking_context_from_analysis(state: ExperimentSuggestionState) -> str:
    """Create enhanced ranking context for experiment suggestions using extracted analysis."""
    # Start with the original user query
    context_parts = [f"User Query: {state['original_prompt']}"]
    
    # Add findings analysis if available
    findings_analysis = state.get("findings_analysis", {})
    if findings_analysis:
        # Add domain information
        domain_analysis = findings_analysis.get("domain_analysis", {})
        if domain_analysis:
            context_parts.append("Research Domain Context:")
            if domain_analysis.get("primary_domain"):
                context_parts.append(f"- Primary Domain: {domain_analysis['primary_domain']}")
            if domain_analysis.get("task_type"):
                context_parts.append(f"- Task Type: {domain_analysis['task_type']}")
            if domain_analysis.get("application_area"):
                context_parts.append(f"- Application: {domain_analysis['application_area']}")
            if domain_analysis.get("data_type"):
                context_parts.append(f"- Data Type: {domain_analysis['data_type']}")
        
        # Add research opportunities 
        opportunities = findings_analysis.get("research_opportunities", [])
        if opportunities:
            context_parts.append("Research Focus Areas:")
            for opp in opportunities[:3]:  # Top 3 opportunities
                context_parts.append(f"- {opp}")
        
        # Add current state information
        current_state = findings_analysis.get("current_state", {})
        if current_state and current_state.get("findings"):
            context_parts.append(f"Current Research State: {current_state['findings']}")
    
    # Add research direction if available
    research_direction = state.get("research_direction", {})
    if research_direction:
        selected_direction = research_direction.get("selected_direction", {})
        if selected_direction.get("direction"):
            context_parts.append(f"Research Direction: {selected_direction['direction']}")
        
        # Add key questions
        key_questions = selected_direction.get("key_questions", [])
        if key_questions:
            context_parts.append("Key Research Questions:")
            for question in key_questions[:2]:  # Top 2 questions
                context_parts.append(f"- {question}")
    
    # Combine all parts
    ranking_context = '\n'.join(context_parts)
    
    # Limit total length to avoid token issues
    if len(ranking_context) > 1500:
        ranking_context = ranking_context[:1500] + "..."
    
    return ranking_context


def create_custom_ranking_prompt(prompt_type: str = "default") -> str:
    """Create a custom ranking prompt based on prompt type."""
    
    if prompt_type == "experimental":
        return """
            You are an expert experimental methodology researcher.  
            Your task: Estimate how relevant this paper is to **experimental research needs** using ONLY the paper’s title and summary (abstract).  

            OUTPUT FORMAT (STRICT):
            - Return EXACTLY one floating-point number with ONE decimal (e.g., 8.7).  
            - No words, no JSON, no units, no symbols, no explanation.  
            - Single line only (no leading/trailing spaces or extra lines).  

            SCORING CRITERIA (use inference from title/summary):  
            - methodology_relevance (40%): Does the summary explicitly mention experimental methodology, benchmarks, protocols, or evaluation setups?  
            - experimental_evidence (30%): Does it mention results, experiments, performance comparisons, or ablation studies?  
            - implementation_guidance (20%): Does it provide or strongly imply practical details like datasets, code availability, reproducibility, or implementation notes?  
            - research_alignment (10%): Does it align with the given research direction and questions?  

            COMPUTE:  
            - Let m,e,i,r ∈ [0,1], estimated from the title/summary.  
            - score = round((0.40*m + 0.30*e + 0.20*i + 0.10*r) * 10, 1).  
            - If the title/summary clearly lacks experimental content (all four < 0.15), output **1.0**.  
            - Clip final result to [1.0, 10.0].  

            PRIORITIZATION:  
            - Favor papers with explicit mention of **empirical studies, benchmarks, datasets, or evaluation frameworks**.  
            - Penalize papers that are purely theoretical, conceptual, or survey-style with no experimental grounding.  
            Research context:
            \"\"\"{query}\"\"\"

            Paper title:
            \"\"\"{title}\"\"\"

            Paper summary:
            \"\"\"{content}\"\"\"
        """.strip()
    
    elif prompt_type == "model_suggestion":
        return """
            You are an expert ML model selection researcher. Score how relevant this paper is to model selection and architecture research on a 1–10 scale.

            OUTPUT FORMAT (STRICT):
            - Return EXACTLY one floating-point number with ONE decimal (e.g., 8.7).
            - No words, no JSON, no units, no symbols, no explanation.
            - Single line only (no leading/trailing spaces or extra lines).

            MODEL FOCUS SCORING - assign four subscores in [0,1]:
            - architecture_relevance (40%): discusses relevant model architectures, neural network designs, or ML approaches
            - performance_evidence (30%): provides performance benchmarks, comparisons, or evaluation results
            - implementation_details (20%): includes implementation specifics, hyperparameters, training procedures, or code
            - task_alignment (10%): addresses similar tasks, domains, or application requirements

            Compute:
            - Let a,p,i,t ∈ [0,1].
            - score = round((0.40*a + 0.30*p + 0.20*i + 0.10*t) * 10, 1).
            - If clearly unrelated to models/architectures (all four < 0.15), output 1.0.
            - Clip to [1.0, 10.0].
            - Prioritize papers with concrete model architectures and performance data.

            Research context:
            \"\"\"{query}\"\"\"

            Paper title:
            \"\"\"{title}\"\"\"

            Paper summary:
            \"\"\"{content}\"\"\"
        """.strip()
                    
    else:  # default prompt
        return None  # Use the original prompt in arxiv_paper_utils.py


# ==================================================================================
# WORKFLOW NODES
# ==================================================================================

async def _analyze_experiment_findings_node(state: ExperimentSuggestionState) -> ExperimentSuggestionState:
    """Node for analyzing experimental findings and research context for experiment suggestions."""
    _write_stream("Experiment Analysis: Analyzing current findings and research context...")
    
    # Extract dependencies from state (adapting from class method)
    client = state["client"]
    model = state["model"]
    
    try:
        # Extract research context from the original prompt and any provided data
        original_prompt = state.get("original_prompt", "")
        uploaded_data = state.get("uploaded_data", [])
        experimental_results = state.get("experimental_results", {})
        
        # Combine user query with uploaded data (CSV or other file contents)
        full_prompt_context = _combine_query_and_data(original_prompt, uploaded_data)
        
        
        # Display uploaded data info if present
       
        # Check for previous analysis iterations and validation feedback
        analysis_iterations = state.get("analysis_iterations", [])
        analysis_validation_results = state.get("analysis_validation_results", {})
        current_iteration = len(analysis_iterations) + 1
        
        # DEBUG: Detailed iteration tracking
        #print(f"🐛 DEBUG _analyze_experiment_findings_node:")
        #print(f"  analysis_iterations length: {len(analysis_iterations)}")
        #print(f"  current_iteration calculated: {current_iteration}")
        #if analysis_iterations:
         #   #print(f"  last iteration in history: {analysis_iterations[-1].get('iteration', 'NO_ITERATION')}")
        #print(f"  state keys: {sorted(state.keys())}")

        # Extract validation feedback for improvement
        validation_feedback = ""
        if analysis_validation_results and current_iteration > 1:
            critical_issues = analysis_validation_results.get("critical_issues", [])
            completeness_gaps = analysis_validation_results.get("completeness_gaps", [])
            accuracy_concerns = analysis_validation_results.get("accuracy_concerns", [])
            improvement_recommendations = analysis_validation_results.get("improvement_recommendations", [])
            
            validation_feedback = f"""
                PREVIOUS VALIDATION FEEDBACK (Iteration {current_iteration - 1}):

                Critical Issues to Address:
                {chr(10).join(f"• {issue}" for issue in critical_issues[:3])}

                Completeness Gaps to Fill:
                {chr(10).join(f"• {gap}" for gap in completeness_gaps[:3])}

                Accuracy Concerns to Resolve:
                {chr(10).join(f"• {concern}" for concern in accuracy_concerns[:3])}

                Improvement Recommendations:
                {chr(10).join(f"• {rec}" for rec in improvement_recommendations[:3])}

                CRITICAL: Address ALL the above issues in this iteration. Provide specific, accurate, and complete analysis.
            """

        _write_stream(f"Generating analysis (iteration {current_iteration})")
        if validation_feedback:
            _write_stream(f"Incorporating validation feedback from previous iteration")

        # Enhanced analysis prompt for experiment suggestions
        iteration_context = f"ITERATION {current_iteration}" if current_iteration > 1 else "INITIAL ANALYSIS"
        
        analysis_prompt = f"""
        You are an expert machine learning researcher analyzing experimental findings to suggest follow-up experiments.
        
        **{iteration_context}**
        {validation_feedback}
        
        Original Research Question/Problem: "{original_prompt}"
        
        Experimental Results/Context: {experimental_results if experimental_results else "User described their current experimental situation in the prompt above"}
        
        Uploaded data: {uploaded_data if uploaded_data else "No additional data provided"}
        
        **CRITICAL**: If the user hasn't explicitly described their problem domain, you must infer it from:
        - Keywords in their prompt (computer vision, NLP, object detection, classification, etc.)
        - Data types mentioned (images, text, sensor data, time series, etc.)
        - Models or techniques referenced (CNNs, transformers, YOLO, etc.)
        - Applications mentioned (autonomous driving, medical imaging, etc.)
        
        VALIDATION REQUIREMENTS FOR ANALYSIS ACCURACY:
        1. **Be SPECIFIC and TECHNICAL** - avoid generic observations
        2. **Provide ACTIONABLE insights** - researchers must be able to act on your analysis
        3. **Ground analysis in DOMAIN EXPERTISE** - demonstrate deep understanding of the field
        4. **Include PRECISE technical details** - methods, datasets, metrics, benchmarks
        5. **Address validation feedback** (if provided above) - fix all identified issues
        6. **Ensure COMPLETENESS** - cover all required analysis sections thoroughly
        
        Please analyze this research context and provide a comprehensive analysis for experiment planning:
        
        1. **Domain Inference and Analysis** (BE SPECIFIC):
           - Primary research domain (computer vision, NLP, reinforcement learning, robotics, etc.)
           - Specific task type (object detection, classification, segmentation, generation, etc.)
           - Application area (autonomous vehicles, medical imaging, robotics, etc.)
           - Data characteristics (images, text, sensor data, time series, tabular, etc.)
           - Technical complexity level and requirements
           
        2. **Current State Assessment** (PROVIDE DETAILS):
           - What has been accomplished so far?
           - What are the key findings or results?
           - What metrics or performance indicators are being used?
           - Current model architectures or approaches being used
           - Performance baseline and targets
           
        3. **Technical Context** (INCLUDE SPECIFICS):
           - Frameworks and methodologies currently employed
           - Datasets being used or dataset characteristics
           - Computational constraints or requirements
           - Evaluation benchmarks and metrics
           - Hardware and software requirements
           
        4. **Research Gaps and Opportunities** (BE ACTIONABLE):
           - What questions remain unanswered in this domain?
           - What aspects need deeper investigation?
           - What are common failure modes or limitations in this area?
           - Areas for improvement or optimization
           - Specific research directions to pursue
           
        5. **Domain-Specific Considerations** (DEMONSTRATE EXPERTISE):
           - Key challenges specific to this research domain
           - Standard experimental practices in this field
           - Important datasets, benchmarks, or evaluation protocols
           - State-of-the-art methods and their limitations
           - Future research trends and opportunities
          
        
        Return your analysis in JSON format with clear, domain-specific insights that will inform targeted literature search and experiment suggestions.
        
        Example for computer vision/object detection:
        {{
            "domain_analysis": {{
                "primary_domain": "computer vision",
                "task_type": "object detection",
                "application_area": "autonomous driving",
                "data_type": "images/video"
            }},
            "current_state": {{
                "findings": "Model achieving X mAP on dataset Y",
                "current_approach": "YOLO-based detection pipeline"
            }},
            "research_opportunities": [
                "Multi-scale detection improvements",
                "Real-time inference optimization",
                "Small object detection enhancement"
            ]
        }}
        """
        
        
        response = client.chat.completions.create(
            model=model,
            temperature=0.2,
            messages=[{"role": "user", "content": analysis_prompt}],
        )

        
        # Parse the analysis
        analysis_text = response.choices[0].message.content.strip()
        #print(f"📋 Research Analysis: {analysis_text}")
        
        # Try to extract JSON from response
        try:
            # Clean and extract JSON from response
            if analysis_text.startswith("```json"):
                analysis_text = analysis_text[7:]
            if analysis_text.endswith("```"):
                analysis_text = analysis_text[:-3]
            analysis_text = analysis_text.strip()
            
            # Look for JSON in the response
            start = analysis_text.find('{')
            end = analysis_text.rfind('}') + 1
            if start != -1 and end != -1:
                analysis_json = json.loads(analysis_text[start:end])
                #print(f"✅ Successfully parsed research analysis JSON")
                
                # Validate and ensure required fields are present
                if "research_opportunities" not in analysis_json:
                    analysis_json["research_opportunities"] = ["Experimental validation", "Comparative studies", "Performance optimization"]
                    #print("⚠️ Added missing research_opportunities field")
                
                if "current_state" not in analysis_json:
                    analysis_json["current_state"] = {"status": "Analysis from LLM", "findings": "Based on user prompt and analysis"}
                    #print("⚠️ Added missing current_state field")
                    
            else:
                # Fallback: create structured analysis from domain inference
                #print("No JSON found, creating structured analysis...")
                
                # Try to infer domain from original prompt
                prompt_lower = original_prompt.lower()
                domain_info = _infer_domain_from_prompt(prompt_lower)
                
                analysis_json = {
                    "domain_analysis": domain_info,
                    "current_state": {"status": "Initial analysis", "findings": "Based on user prompt"},
                    "research_opportunities": ["Experimental validation", "Comparative studies", "Performance optimization"],
                    "summary": "Analysis based on prompt content and domain inference"
                }
                
        except json.JSONDecodeError as e:
            _write_stream(f"JSON parsing failed: {e}, creating fallback analysis...")
            # Fallback: create structured analysis
            prompt_lower = original_prompt.lower()
            domain_info = _infer_domain_from_prompt(prompt_lower)
            
            analysis_json = {
                "domain_analysis": domain_info,
                "current_state": {"status": "Initial analysis", "findings": "Based on user prompt"},
                "research_opportunities": ["Experimental validation", "Comparative studies", "Performance optimization"],
                "summary": f"Research analysis for {domain_info.get('primary_domain', 'machine learning')} project"
            }
                
        except Exception as e:
            _write_stream(f"JSON parsing failed: {e}, using fallback analysis")
            # Fallback analysis with extracted key information
            prompt_lower = original_prompt.lower()
            domain_info = _infer_domain_from_prompt(prompt_lower)
            
            analysis_json = {
                "domain_analysis": domain_info,
                "current_state": {"status": "Analysis from prompt", "findings": original_prompt[:200]},
                "research_opportunities": ["Follow-up experiments", "Comparative analysis"],
                "summary": f"Fallback analysis for {domain_info.get('primary_domain', 'machine learning')} research"
            }
        
        # Store the analysis and update state
        return {
            **state,
            "findings_analysis": analysis_json,
            "current_analysis_iteration": current_iteration,
            "research_context": {
                "original_prompt": original_prompt,
                "domain": analysis_json.get("domain_analysis", {}).get("primary_domain", "machine learning"),
                "analysis_timestamp": time.strftime("%Y-%m-%d %H:%M:%S")
            },
            "analysis_completed": True,
            "current_step": "findings_analyzed"
        }
        
    except Exception as e:
        print(f"Error in analyze_experiment_findings: {str(e)}")
        return {
            **state,
            "current_analysis_iteration": state.get("current_analysis_iteration", 1),
            "errors": state.get("errors", []) + [f"Findings analysis error: {str(e)}"],
            "analysis_completed": False,
            "current_step": "analysis_error"
        }


def _infer_domain_from_prompt(prompt_lower: str) -> dict:
    """Infer research domain from user prompt keywords."""
    domain_info = {
        "primary_domain": "machine learning",
        "task_type": "experimental",
        "application_area": "",
        "data_type": ""
    }
    
    # Computer Vision keywords
    if any(kw in prompt_lower for kw in ["computer vision", "cv", "image", "video", "visual", "detection", "segmentation", "yolo", "cnn", "resnet"]):
        domain_info["primary_domain"] = "computer vision"
        if "detection" in prompt_lower:
            domain_info["task_type"] = "object detection"
        elif "segmentation" in prompt_lower:
            domain_info["task_type"] = "segmentation"
        elif "classification" in prompt_lower and "image" in prompt_lower:
            domain_info["task_type"] = "image classification"
        domain_info["data_type"] = "images"
        
    # NLP keywords
    elif any(kw in prompt_lower for kw in ["nlp", "text", "language", "transformer", "bert", "gpt", "sentiment", "translation"]):
        domain_info["primary_domain"] = "NLP"
        if "classification" in prompt_lower:
            domain_info["task_type"] = "text classification"
        elif "generation" in prompt_lower:
            domain_info["task_type"] = "text generation"
        elif "translation" in prompt_lower:
            domain_info["task_type"] = "machine translation"
        domain_info["data_type"] = "text"
        
    # Robotics keywords
    elif any(kw in prompt_lower for kw in ["robot", "autonomous", "control", "navigation", "manipulation"]):
        domain_info["primary_domain"] = "robotics"
        domain_info["task_type"] = "control"
        domain_info["data_type"] = "sensor data"
        
    # Time series keywords
    elif any(kw in prompt_lower for kw in ["time series", "temporal", "forecasting", "lstm", "sequence"]):
        domain_info["primary_domain"] = "time series analysis"
        domain_info["task_type"] = "forecasting"
        domain_info["data_type"] = "time series"
        
    # Application areas
    if any(kw in prompt_lower for kw in ["medical", "healthcare", "clinical"]):
        domain_info["application_area"] = "medical"
    elif any(kw in prompt_lower for kw in ["autonomous", "driving", "vehicle"]):
        domain_info["application_area"] = "autonomous driving"
    elif any(kw in prompt_lower for kw in ["finance", "financial", "trading"]):
        domain_info["application_area"] = "finance"
        
    return domain_info

async def _validate_analysis_node(state: ExperimentSuggestionState) -> ExperimentSuggestionState:
    """Node for validating the generated data analysis with hyper-strict criteria."""
    client = state["client"]
    model = state["model"]
    _write_stream("Analysis Validation: Evaluating generated data analysis for accuracy, completeness, and grounding.")
    
    try:
        # Extract current analysis and context
        findings_analysis = state.get("findings_analysis", {})
        original_prompt = state.get("original_prompt", "")
        experimental_results = state.get("experimental_results", {})
        uploaded_data = state.get("uploaded_data", [])
        
        # Track analysis iteration history
        analysis_iterations = state.get("analysis_iterations", [])
        current_iteration = len(analysis_iterations) + 1
        
        # DEBUG: Track validation iteration calculation
        #print(f"🐛 DEBUG _validate_analysis_node:")
        ##print(f"  analysis_iterations length: {len(analysis_iterations)}")
        #print(f"  current_iteration calculated: {current_iteration}")
        #if analysis_iterations:
        #    print(f"  last iteration in history: {analysis_iterations[-1].get('iteration', 'NO_ITERATION')}")
        
        # Add current analysis to history
        current_analysis_record = {
            "iteration": current_iteration,
            "analysis": str(findings_analysis)[:500] if findings_analysis else "No analysis generated",
            "timestamp": __import__('datetime').datetime.now().isoformat()
        }
        analysis_iterations.append(current_analysis_record)
        
        # Create iteration history context
        iteration_history = ""
        if len(analysis_iterations) > 1:
            iteration_history = "\n\nPREVIOUS ANALYSIS ITERATIONS:\n"
            for i, iteration in enumerate(analysis_iterations[:-1], 1):
                iteration_history += f"Iteration {iteration['iteration']}: {iteration['analysis'][:100]}...\n\n"
        
        validation_prompt = f"""
            You are a HYPER-STRICT data analysis validator. Your job is to rigorously evaluate the generated analysis for **accuracy**, **completeness**, **logical consistency**, **domain expertise**, and **actionable insights**.

            ORIGINAL RESEARCH REQUEST:
            {original_prompt}

            EXPERIMENTAL CONTEXT:
            {experimental_results}
            
            Uploaded data: 
            {uploaded_data if uploaded_data else "No additional data provided"}

            GENERATED ANALYSIS:
            {findings_analysis}

            CURRENT ITERATION: {current_iteration}
            {iteration_history}

            HYPER-STRICT VALIDATION CRITERIA (ALL MUST BE SATISFIED):
            1. **Domain Accuracy**: Is the domain identification and characterization correct and specific?
            2. **Technical Completeness**: Does the analysis include all necessary technical details (datasets, methods, metrics)?
            3. **Logical Consistency**: Are all conclusions logically supported by the provided context?
            4. **Insight Quality**: Does the analysis provide actionable, meaningful insights beyond obvious observations?
            5. **Contextual Grounding**: Is the analysis properly grounded in the user's specific research context?
            6. **Gap Identification**: Are research gaps and opportunities clearly and accurately identified?
            7. **Technical Depth**: Does the analysis demonstrate appropriate domain expertise and technical understanding?
            8. **Actionability**: Can researchers actually use this analysis to make informed decisions?
            9. **Specificity**: Are recommendations specific enough to be implementable rather than vague?
            10. **Accuracy Verification**: Are all technical claims and domain assertions verifiable and correct?

            ZERO-TOLERANCE REQUIREMENTS:
            - No generic or template-like responses
            - No vague or aspirational language without specifics
            - No technical inaccuracies or domain mischaracterizations
            - No missing critical analysis components
            - No unsupported claims or assumptions

            ITERATION ANALYSIS:
            - If this is iteration 1: Apply HYPER-STRICT validation
            - If iteration 2+: Ensure ALL previous issues resolved AND significant improvement demonstrated

            Return your assessment in this exact JSON format:
            {{
                "validation_result": "PASS" | "FAIL",
                "overall_score": 0.0-1.0,
                "detailed_scores": {{
                    "domain_accuracy": 0.0-1.0,
                    "technical_completeness": 0.0-1.0,
                    "logical_consistency": 0.0-1.0,
                    "insight_quality": 0.0-1.0,
                    "contextual_grounding": 0.0-1.0,
                    "gap_identification": 0.0-1.0,
                    "technical_depth": 0.0-1.0,
                    "actionability": 0.0-1.0,
                    "specificity": 0.0-1.0,
                    "accuracy_verification": 0.0-1.0
                }},
                "critical_issues": ["list", "of", "critical", "problems"],
                "completeness_gaps": ["missing", "analysis", "components"],
                "accuracy_concerns": ["technical", "inaccuracies", "or", "concerns"],
                "improvement_recommendations": ["specific", "actionable", "improvements"],
                "decision_rationale": "Clear explanation of pass/fail decision focusing on accuracy and completeness",
                "iteration_assessment": "Analysis of improvement from previous iterations (if applicable)",
                "confidence_in_validation": 0.0-1.0
            }}

            RUTHLESS PASSING THRESHOLD: Overall score ≥ 0.90 AND no critical issues AND no completeness gaps AND no accuracy concerns AND all detailed scores ≥ 0.85
            BE ABSOLUTELY STRICT: Only pass analyses that are **technically perfect**, **completely accurate**, **deeply insightful**, and **fully actionable**.
        """

        # Call LLM for validation
        
        
        response = client.chat.completions.create(
            model=model,
            temperature=0.2,
            messages=[ {"role": "system", "content": "You are a ruthlessly strict data analysis validator. Provide objective, rigorous assessments in valid JSON format. Be ultra-conservative - only pass analyses that are technically perfect and deeply insightful."},
                    {"role": "user", "content": validation_prompt}]
        )

        validation_content = response.choices[0].message.content.strip()
        
        # Parse validation response
        try:
            # Clean and extract JSON
            json_match = re.search(r'\{.*\}', validation_content, re.DOTALL)
            if json_match:
                validation_json = json.loads(json_match.group(0))
            else:
                raise json.JSONDecodeError("No JSON found", validation_content, 0)
            
            validation_result = validation_json.get("validation_result", "FAIL").upper()
            overall_score = validation_json.get("overall_score", 0.0)
            critical_issues = validation_json.get("critical_issues", [])
            completeness_gaps = validation_json.get("completeness_gaps", [])
            accuracy_concerns = validation_json.get("accuracy_concerns", [])
            improvement_recommendations = validation_json.get("improvement_recommendations", [])
            _write_stream(f"Analysis Validation Result: {validation_result} with overall score {overall_score:.2f}, needs 0.9 to pass.")
            _write_stream(f"Critical Issues: {len(critical_issues)}, Completeness Gaps: {len(completeness_gaps)}, Accuracy Concerns: {len(accuracy_concerns)}")
            # Safety check: Enforce RUTHLESS thresholds
            if overall_score < 0.90 or len(critical_issues) > 0 or len(completeness_gaps) > 0 or len(accuracy_concerns) > 0:
                validation_result = "FAIL"
            
            # Check iteration limit (max 3 iterations to prevent infinite loops)
            if current_iteration >= 3 and validation_result == "FAIL":
                _write_stream(f"⚠️ Maximum analysis iterations reached ({current_iteration}). Forcing continuation with current analysis with {len(critical_issues)}, critical issues, {len(completeness_gaps)} completeness gaps, {len(accuracy_concerns)} accuracy concerns.")
                #print(f"🚨 WARNING: Validation found {len(critical_issues)} critical issues, {len(completeness_gaps)} completeness gaps, {len(accuracy_concerns)} accuracy concerns.")
                #print(f"🚨 This is a FORCED PASS to prevent infinite loops - analysis has unresolved validation issues!")
                validation_result = "PASS"
                validation_json["forced_pass"] = True
                validation_json["decision_rationale"] = f"Forced pass after {current_iteration} iterations to prevent infinite loop. Original validation failed due to: {len(critical_issues)} critical issues, {len(completeness_gaps)} completeness gaps, {len(accuracy_concerns)} accuracy concerns."
            '''
            print("\n" + "=" * 80)
            print("🔍 HYPER-STRICT ANALYSIS VALIDATION RESULTS")
            print("=" * 80)
            print(f"📊 Iteration: {current_iteration}")
            if validation_json.get("forced_pass"):
                print(f"🎯 Validation Result: {validation_result} (⚠️ FORCED PASS - VALIDATION FAILED)")
            else:
                print(f"🎯 Validation Result: {validation_result}")
            print(f"📈 Overall Score: {overall_score:.2f}/1.0 (Required: ≥0.90)")
            print(f"🔴 Critical Issues: {len(critical_issues)}")
            print(f"📋 Completeness Gaps: {len(completeness_gaps)}")
            print(f"⚠️ Accuracy Concerns: {len(accuracy_concerns)}")
            
            if critical_issues:
                print("Critical Issues:")
                for issue in critical_issues[:3]:
                    print(f"  • {issue}")
            
            if completeness_gaps:
                print("Completeness Gaps:")
                for gap in completeness_gaps[:3]:
                    print(f"  • {gap}")
            
            if accuracy_concerns:
                print("Accuracy Concerns:")
                for concern in accuracy_concerns[:3]:
                    print(f"  • {concern}")
            
            if validation_result == "FAIL" and improvement_recommendations:
                print("🔧 Improvement Recommendations:")
                for rec in improvement_recommendations[:3]:
                    print(f"  • {rec}")
            
            print(f"💭 Decision Rationale: {validation_json.get('decision_rationale', 'No rationale provided')}")
            print("=" * 80)
            
            # DEBUG: Show what we're returning
            print(f"🐛 DEBUG validation node returning:")
            print(f"  validation_result: {validation_result}")
            print(f"  analysis_validation_decision: {validation_result}")
            print(f"  analysis_iterations length: {len(analysis_iterations)}")
            '''
            # Store validation results in state
            updated_state = {
                **state,
                "analysis_validation_results": validation_json,
                "analysis_iterations": analysis_iterations,
                "analysis_validation_decision": validation_result,
                "current_analysis_iteration": current_iteration,
                "current_step": "analysis_validated"
            }
            
            # DEBUG: Verify the state we're returning
           # print(f"🐛 DEBUG state being returned has keys: {list(updated_state.keys())}")
           # print(f"🐛 DEBUG state['analysis_validation_decision'] = {updated_state.get('analysis_validation_decision')}")
            
            # CRITICAL FIX: Make the routing decision here instead of in separate function
            # Check if this was a forced pass due to max iterations
            forced_pass = validation_json.get("forced_pass", False)
            
            # Safety check: After 3 iterations, force continue to avoid infinite loops
            if current_iteration >= 3:
                if forced_pass:
                    _write_stream(f"Maximum analysis iterations reached ({current_iteration}). Forced pass due to iteration limit - continuing to research direction despite validation issues.")
                else:
                    _write_stream(f"Maximum analysis iterations reached ({current_iteration}). Continuing to research direction.")
                updated_state["next_node"] = "decide_research_direction"
            # Check validation result - but distinguish between genuine pass and forced pass
            elif validation_result == "PASS":
                if forced_pass:
                    _write_stream(f"Analysis validation was FORCED to pass after max iterations. Continuing to research direction with unresolved issues.")
                else:
                    _write_stream(f"Analysis validation passed. Continuing to research direction.")
                updated_state["next_node"] = "decide_research_direction"
            else:
                _write_stream(f"Analysis validation failed. Iterating to improve analysis (iteration {current_iteration + 1}).")
                updated_state["next_node"] = "analyze_findings"
            
            return updated_state
            
        except json.JSONDecodeError as e:
            print(f"❌ Failed to parse analysis validation JSON: {e}")
            # Default to FAIL for safety
            fallback_validation = {
                "validation_result": "FAIL",
                "overall_score": 0.4,
                "critical_issues": ["JSON parsing error in analysis validation"],
                "improvement_recommendations": ["Regenerate analysis with clearer structure"],
                "decision_rationale": "Analysis validation failed due to parsing error",
                "error": str(e)
            }
            
            return {
                **state,
                "analysis_validation_results": fallback_validation,
                "analysis_iterations": analysis_iterations,
                "analysis_validation_decision": "FAIL",
                "current_analysis_iteration": current_iteration,
                "current_step": "analysis_validation_error",
                "next_node": "analyze_findings"  # Always retry on error
            }
            
    except Exception as e:
        print(f"Error in validate_analysis: {str(e)}")
        # Default to PASS to avoid blocking the workflow
        error_validation = {
            "validation_result": "PASS",
            "overall_score": 0.6,
            "decision_rationale": f"Analysis validation error occurred: {str(e)}. Defaulting to PASS to continue workflow.",
            "error": str(e)
        }
        
        return {
            **state,
            "analysis_validation_results": error_validation,
            "analysis_iterations": analysis_iterations,
            "analysis_validation_decision": "FAIL",
            "current_analysis_iteration": current_iteration,
            "errors": state.get("errors", []) + [f"Analysis validation error: {str(e)}"],
            "current_step": "analysis_validation_error_pass",
            "next_node": "analyze_findings"  # Always retry on error
        }



async def _decide_research_direction_node(state: ExperimentSuggestionState) -> ExperimentSuggestionState:
    """Node for deciding the research direction based on analysis findings."""
    _write_stream("Research Direction: Determining optimal research path with justification.")
    client = state["client"]
    model = state["model"]
    try:
        # Extract analysis context
        original_prompt = state.get("original_prompt", "")
        findings_analysis = state.get("findings_analysis", {})
        experimental_results = state.get("experimental_results", {})
        research_context = state.get("research_context", {})
        
        # Extract iteration history and validation feedback
        direction_iterations = state.get("direction_iterations", [])
        validation_results = state.get("direction_validation_results", {})
        current_iteration = len(direction_iterations) + 1
        
        # Prepare context for direction decision
        findings_summary = findings_analysis.get("summary", "No analysis available")
        key_insights = findings_analysis.get("key_insights", [])
        limitations = findings_analysis.get("limitations", [])
        
        # Create iteration history context
        iteration_context = ""
        validation_feedback = ""
        
        if current_iteration > 1:
            iteration_context = "\n\nPREVIOUS RESEARCH DIRECTION ATTEMPTS:\n"
            for i, iteration in enumerate(direction_iterations, 1):
                iteration_context += f"Attempt {iteration['iteration']}: {iteration['direction']}\n"
                iteration_context += f"  Issues: Failed validation\n"
                iteration_context += f"  Confidence: {iteration['confidence_level']}\n\n"
            
            if validation_results.get("critical_issues"):
                validation_feedback = f"\n\nCRITICAL ISSUES FROM PREVIOUS VALIDATION:\n"
                for issue in validation_results.get("critical_issues", [])[:3]:
                    validation_feedback += f"• {issue}\n"
            
            if validation_results.get("improvement_recommendations"):
                validation_feedback += f"\nIMPROVEMENT RECOMMENDATIONS:\n"
                for rec in validation_results.get("improvement_recommendations", [])[:3]:
                    validation_feedback += f"• {rec}\n"
        
        # Create comprehensive prompt for direction decision
        direction_prompt = f"""
        You are a senior research strategist. Based on the experimental findings and analysis, determine the most promising research direction to pursue next.

        ORIGINAL RESEARCH CONTEXT:
        {original_prompt}

        EXPERIMENTAL FINDINGS SUMMARY:
        {findings_analysis if findings_analysis else "No analysis available"}

        KEY INSIGHTS FROM ANALYSIS:
        {chr(10).join(f"• {insight}" for insight in key_insights[:5])}

        IDENTIFIED LIMITATIONS:
        {chr(10).join(f"• {limitation}" for limitation in limitations[:3])}

        EXPERIMENTAL DATA OVERVIEW:
        {str(experimental_results)[:500] if experimental_results else "No experimental data provided"}

        ITERATION CONTEXT:
        Current Iteration: {current_iteration}
        {iteration_context}
        {validation_feedback}

        INSTRUCTIONS FOR ITERATION {current_iteration}:
        {"INITIAL DIRECTION GENERATION:" if current_iteration == 1 else f"IMPROVED DIRECTION GENERATION (addressing previous failures):"}

        Your task is to:

        - Identify 2 novel future research directions that go beyond the current study’s scope, inspired by its findings but not repeating them.

        - Select the most promising direction and provide a clear justification for this choice.

        - Explain why this direction is optimal given the strengths and limitations of the current results.

        - Outline the specific aspects, variables, or methodologies that should be investigated to pursue this new direction.
        
        {"CRITICAL FOR ITERATION " + str(current_iteration) + ": The previous direction failed validation. You MUST significantly improve by addressing validation feedback and avoiding previous issues." if current_iteration > 1 else ""}
        
        Return your response in this exact JSON format:
        {{
            "potential_directions": [
                {{
                    "direction": "Brief direction description",
                    "rationale": "Why this direction is promising",
                    "feasibility": "Assessment of feasibility (High/Medium/Low)"
                }}
            ],
            "selected_direction": {{
                "direction": "Chosen research direction",
                "justification": "Detailed explanation of why this direction was selected",
                "expected_impact": "What outcomes this direction could achieve",
                "key_questions": ["Question 1", "Question 2", "Question 3"],
                "confidence_level": "High/Medium/Low"
            }},
            "reasoning": "Overall strategic reasoning for the decision"
        }}
        """

        # Call LLM for direction decision
        response = client.chat.completions.create(
            model=model,
            temperature=0.2,
            messages=[{"content": direction_prompt, "role": "user"}]
        )
        
        direction_content = response.choices[0].message.content.strip()
        
        # Clean and parse JSON response
        import json
        import re
        
        # Clean the response to extract JSON
        json_match = re.search(r'\{.*\}', direction_content, re.DOTALL)
        if json_match:
            clean_json = json_match.group()
            try:
                direction_decision = json.loads(clean_json)
            except json.JSONDecodeError:
                # Fallback if JSON parsing fails
                direction_decision = {
                    "selected_direction": {
                        "direction": "Continue current research with refinements",
                        "justification": "Based on analysis, refining current approach shows promise",
                        "expected_impact": "Improved understanding of the research problem",
                        "key_questions": ["How to optimize current methods?", "What are the key bottlenecks?", "What alternative approaches exist?"],
                        "confidence_level": "Medium"
                    },
                    "reasoning": "Default direction based on analysis findings"
                }
        else:
            # Fallback direction
            direction_decision = {
                "selected_direction": {
                    "direction": "Investigate identified limitations and explore alternatives",
                    "justification": "Analysis revealed limitations that need addressing",
                    "expected_impact": "Better understanding of constraints and potential solutions",
                    "key_questions": ["What causes the identified limitations?", "What alternative methods exist?", "How can we validate improvements?"],
                    "confidence_level": "Medium"
                },
                "reasoning": "Focus on addressing key limitations identified in analysis"
            }

        _write_stream(f"Research direction decided: {direction_decision.get('selected_direction', {}).get('direction', 'Unknown')}")
        
        return {
            **state,
            "research_direction": direction_decision,
            "current_step": "direction_decided"
        }
        
    except Exception as e:
        print(f"❌ Error in decide_research_direction: {str(e)}")
        # Provide fallback direction
        fallback_direction = {
            "selected_direction": {
                "direction": "Continue systematic investigation",
                "justification": "Maintain research momentum while addressing any identified issues",
                "expected_impact": "Steady progress toward research objectives",
                "key_questions": ["What are the next logical steps?", "How can we improve current methods?", "What additional data is needed?"],
                "confidence_level": "Medium"
            },
            "reasoning": "Fallback direction due to processing error"
        }
        
        return {
            **state,
            "research_direction": fallback_direction,
            "errors": state.get("errors", []) + [f"Direction decision error: {str(e)}"],
            "current_step": "direction_error"
        }


async def _validate_research_direction_node(state: ExperimentSuggestionState) -> ExperimentSuggestionState:
    """Node for validating the proposed research direction and goals with strict evaluation."""
    _write_stream("Research Direction Validation: Evaluating proposed research goals and methodology.")
    model= state["model"]
    client = state["client"]
    try:
        # Extract current research direction and past iterations
        research_direction = state.get("research_direction", {})
        selected_direction = research_direction.get("selected_direction", {})
        original_prompt = state.get("original_prompt", "")
        findings_analysis = state.get("findings_analysis", {})
        
        # Track iteration history
        direction_iterations = state.get("direction_iterations", [])
        current_iteration = len(direction_iterations) + 1
        
        # Add current direction to history
        current_direction_record = {
            "iteration": current_iteration,
            "direction": selected_direction.get("direction", ""),
            "justification": selected_direction.get("justification", ""),
            "key_questions": selected_direction.get("key_questions", []),
            "confidence_level": selected_direction.get("confidence_level", "Medium"),
            "timestamp": __import__('datetime').datetime.now().isoformat()
        }
        direction_iterations.append(current_direction_record)
        
        # Create validation prompt with iteration history
        iteration_history = ""
        if len(direction_iterations) > 1:
            iteration_history = "\n\nPREVIOUS ITERATION HISTORY:\n"
            for i, iteration in enumerate(direction_iterations[:-1], 1):
                iteration_history += f"Iteration {iteration['iteration']}: {iteration['direction']}\n"
                iteration_history += f"  Confidence: {iteration['confidence_level']}\n"
                iteration_history += f"  Key Questions: {', '.join(iteration['key_questions'][:2])}\n\n"
        
        direction_text = selected_direction.get("direction", "")
        justification = selected_direction.get("justification", "")
        key_questions = selected_direction.get("key_questions", [])
        confidence_level = selected_direction.get("confidence_level", "Medium")
        
        validation_prompt = f"""
            You are a strict research methodology validator. Your job is to rigorously evaluate the proposed research direction for both **scientific soundness** and **strategic value**. You must decide not only if the direction is valid, but also if it is a *worthwhile and meaningful path to pursue* given the research context.

            ORIGINAL RESEARCH REQUEST:
            {original_prompt}

            PROPOSED RESEARCH DIRECTION:
            Direction: {direction_text}
            Justification: {justification}
            Key Questions: {chr(10).join(f"• {q}" for q in key_questions[:5])}
            Confidence Level: {confidence_level}

            CURRENT ITERATION: {current_iteration}
            {iteration_history}

            RESEARCH CONTEXT:
            {findings_analysis}

            STRICT VALIDATION CRITERIA:
            1. **Alignment**: Does this direction directly address the original research request?
            2. **Scientific Rigor**: Are the research questions testable, well-formulated, and methodologically sound?
            3. **Feasibility**: Can this direction realistically be pursued with typical resources (time, compute, expertise)?
            4. **Novelty & Value**: Does this offer meaningful new insights or open promising lines of inquiry, rather than repeating well-trodden ground?
            5. **Impact Potential**: Could this direction lead to significant contributions, practical applications, or field-shaping results?
            6. **Clarity**: Are the objectives, approach, and rationale clearly and coherently defined?
            7. **Scope**: Is the scope balanced (not too broad to be vague, not too narrow to be trivial)?
            8. **Strategic Fit**: Given the research context, is this a *good direction to pursue now* compared to alternatives?

            ITERATION ANALYSIS:
            - If this is iteration 1: Apply standard validation
            - If iteration 2+: Ensure improvements over previous attempts, avoid repeated mistakes

            Return your assessment in this exact JSON format:
            {{
            "validation_result": "PASS" | "FAIL",
            "overall_score": 0.0-1.0,
            "detailed_scores": {{
                "alignment": 0.0-1.0,
                "scientific_rigor": 0.0-1.0,
                "feasibility": 0.0-1.0,
                "novelty_value": 0.0-1.0,
                "impact_potential": 0.0-1.0,
                "clarity": 0.0-1.0,
                "scope": 0.0-1.0,
                "strategic_fit": 0.0-1.0
            }},
            "critical_issues": ["list", "of", "critical", "problems"],
            "improvement_recommendations": ["specific", "actionable", "improvements"],
            "decision_rationale": "Clear explanation of pass/fail decision, weighing novelty, impact, and feasibility",
            "iteration_assessment": "Analysis of improvement from previous iterations (if applicable)",
            "confidence_in_validation": 0.0-1.0
            }}

            PASSING THRESHOLD: Overall score ≥ 0.75 AND no critical issues AND all detailed scores ≥ 0.6
            BE STRICT: Only pass directions that are both **methodologically solid** and **worth pursuing** for novelty and impact.
            """

        # Call LLM for validation
        response = client.chat.completions.create(
            model=model,
            temperature=0.2,
            messages=[{"content": validation_prompt, "role": "user"}]
        )

        validation_content = response.choices[0].message.content.strip()
        
        # Parse validation response
        try:
            # Clean and extract JSON
            json_match = re.search(r'\{.*\}', validation_content, re.DOTALL)
            if json_match:
                validation_json = json.loads(json_match.group(0))
            else:
                raise json.JSONDecodeError("No JSON found", validation_content, 0)
            
            validation_result = validation_json.get("validation_result", "FAIL").upper()
            overall_score = validation_json.get("overall_score", 0.0)
            critical_issues = validation_json.get("critical_issues", [])
            improvement_recommendations = validation_json.get("improvement_recommendations", [])
            
            # Safety check: Enforce strict thresholds
            if overall_score < 0.75 or len(critical_issues) > 0:
                validation_result = "FAIL"
            
            # Check iteration limit (max 3 iterations to prevent infinite loops)
            if current_iteration >= 3 and validation_result == "FAIL":
                _write_stream(f"Maximum iterations reached ({current_iteration}). Forcing continuation with current direction with {len(critical_issues)} critical issues.")
               
               
                validation_result = "PASS"
                validation_json["forced_pass"] = True
                validation_json["decision_rationale"] = f"Forced pass after {current_iteration} iterations to prevent infinite loop. Original validation failed due to: {len(critical_issues)} critical issues."
            ''''
            print("\n" + "=" * 80)
            print("🔍 RESEARCH DIRECTION VALIDATION RESULTS")
            print("=" * 80)
            print(f"📊 Iteration: {current_iteration}")
            if validation_json.get("forced_pass"):
                print(f"🎯 Validation Result: {validation_result} (⚠️ FORCED PASS - VALIDATION FAILED)")
            else:
                print(f"🎯 Validation Result: {validation_result}")
            print(f"📈 Overall Score: {overall_score:.2f}/1.0")
            print(f"🔴 Critical Issues: {len(critical_issues)}")
            hallucination_flags = validation_json.get("hallucination_flags", [])
            
            if critical_issues:
                print("Critical Issues:")
                for issue in critical_issues[:3]:
                    print(f"  • {issue}")
            
            if validation_result == "FAIL" and improvement_recommendations:
                print("🔧 Improvement Recommendations:")
                for rec in improvement_recommendations[:3]:
                    print(f"  • {rec}")
            
            print(f"💭 Decision Rationale: {validation_json.get('decision_rationale', 'No rationale provided')}")
            print("=" * 80)
            '''
           
            # Check if this was a forced pass due to max iterations
            forced_pass = validation_json.get("forced_pass", False)
            
            # Safety check: After 3 iterations, force continue to avoid infinite loops
            if current_iteration >= 3:
                if forced_pass:
                    _write_stream(f"Maximum direction iterations reached ({current_iteration}). Forced pass due to iteration limit - continuing to experiments despite validation issues.")
                else:
                    _write_stream(f"Maximum direction iterations reached ({current_iteration}). Continuing to experiments.")
                next_node = "generate_experiment_search_query"
                
            # Check validation result - but distinguish between genuine pass and forced pass
            elif validation_result == "PASS":
                if forced_pass:
                    _write_stream(f"Research direction validation was FORCED to pass after max iterations. Continuing to experiments with unresolved issues.")
                else:
                    _write_stream(f"Research direction validation passed. Continuing to experiments.")
                next_node = "generate_experiment_search_query"
            else:
                _write_stream(f"Research direction validation failed. Iterating to improve direction (iteration {current_iteration + 1}).")
                next_node = "decide_research_direction"
            
            # Store validation results in state
            return {
                **state,
                "direction_validation_results": validation_json,
                "direction_iterations": direction_iterations,
                "direction_validation_decision": validation_result,
                "current_iteration": current_iteration,
                "current_step": "research_direction_validated",
                "next_node": next_node
            }
            
        except json.JSONDecodeError as e:
            print(f"❌ Failed to parse validation JSON: {e}")
            # Default to FAIL for safety
            fallback_validation = {
                "validation_result": "FAIL",
                "overall_score": 0.4,
                "critical_issues": ["JSON parsing error in validation"],
                "improvement_recommendations": ["Regenerate research direction with clearer objectives"],
                "decision_rationale": "Validation failed due to parsing error",
                "error": str(e)
            }
            
            return {
                **state,
                "direction_validation_results": fallback_validation,
                "direction_iterations": direction_iterations,
                "direction_validation_decision": "FAIL",
                "current_iteration": current_iteration,
                "current_step": "validation_error",
                "next_node": "decide_research_direction"  # Always retry on error
            }
            
    except Exception as e:
        print(f"❌ Error in validate_research_direction: {str(e)}")
        # Default to PASS to avoid blocking the workflow
        error_validation = {
            "validation_result": "PASS",
            "overall_score": 0.6,
            "decision_rationale": f"Validation error occurred: {str(e)}. Defaulting to PASS to continue workflow.",
            "error": str(e)
        }
        
        return {
            **state,
            "direction_validation_results": error_validation,
            "direction_iterations": direction_iterations,
            "direction_validation_decision": "PASS",
            "current_iteration": current_iteration,
            "errors": state.get("errors", []) + [f"Direction validation error: {str(e)}"],
            "current_step": "validation_error_pass",
            "next_node": "generate_experiment_search_query"  # Continue on error with PASS
        }

def _generate_experiment_search_query_node(state: ExperimentSuggestionState) -> ExperimentSuggestionState:
    """Generate ArXiv search query for domain-specific experimental guidance papers."""
    search_iteration = state.get("experiment_search_iteration", 0)
    validation_results = state.get("experiment_paper_validation_results", {})
    is_search_new = validation_results.get("decision") == "search_new"
    
    if search_iteration == 0:
        _write_stream("Experiment Search Query: Generating targeted search for experimental guidance.")
    else:
        _write_stream(f"Experiment Search Query (Retry {search_iteration + 1}): Generating refined search based on validation feedback...")
        
    model = state["model"]
    client = state["client"]
    try:
        # Extract context
        original_prompt = state.get("original_prompt", "")
        research_direction = state.get("research_direction", {})
        findings_analysis = state.get("findings_analysis", {})
        
        selected_direction = research_direction.get("selected_direction", {})
        direction_text = selected_direction.get("direction", "")
        key_questions = selected_direction.get("key_questions", [])
        
        # Extract domain information from analysis
        domain_analysis = findings_analysis.get("domain_analysis", {})
        primary_domain = domain_analysis.get("primary_domain", "machine learning")
        task_type = domain_analysis.get("task_type", "")
        application_area = domain_analysis.get("application_area", "")
        data_type = domain_analysis.get("data_type", "")
        
        # Get previous search query and validation feedback for retries
        previous_query = state.get("experiment_search_query", "")
        search_guidance = {}
        
        if is_search_new and validation_results.get("validation_data"):
            search_guidance = validation_results["validation_data"].get("search_guidance", {})
            _write_stream("Using validation feedback to generate improved search query...")
        
        # Generate domain-specific search query with conditional previous query info
        query_prompt = f"""
        Generate a focused ArXiv search query to find papers in the same research domain that contain experimental methodologies and guidance.

        RESEARCH DOMAIN: {primary_domain}
        TASK TYPE: {task_type}
        APPLICATION: {application_area}
        DATA TYPE: {data_type}
        
        ORIGINAL RESEARCH: {original_prompt}
        
        RESEARCH DIRECTION: {direction_text}
        
        KEY QUESTIONS: {chr(10).join(f"• {q}" for q in key_questions[:3])}"""
        
        # Only add previous query info if it exists and this is a retry
        if previous_query and search_iteration > 0:
            query_prompt += f"""
            
        PREVIOUS FAILED QUERY: {previous_query}
        SEARCH ITERATION: {search_iteration + 1}
        
        VALIDATION FEEDBACK:
        {validation_results.get("reasoning", "Previous papers lacked sufficient experimental methodology")}
        
        SEARCH GUIDANCE FROM VALIDATION:
        - New search terms: {search_guidance.get('new_search_terms', [])}
        - Focus areas: {search_guidance.get('focus_areas', [])}
        - Avoid terms: {search_guidance.get('avoid_terms', [])}
        
        Generate 4 DIFFERENT search terms for ArXiv API, separated by forward slashes (/):
        
        Rules for RETRY query:
        - Term 1: Different primary technique than "{previous_query.split('/')[0] if '/' in previous_query else ''}"
        - Term 2: Alternative task perspective or methodology focus
        - Term 3: Stronger experimental focus (e.g., "experimental validation", "empirical study", "systematic evaluation")
        - Term 4: Specific experimental context (e.g., "ablation", "benchmark", "comparison study")
        
        Focus on finding papers with STRONGER experimental methodology than the previous search.
        Use the suggested new search terms if provided: {search_guidance.get('new_search_terms', [])}"""
        else:
            query_prompt += """

        Generate 4 search terms for ArXiv API, separated by forward slashes (/), that will find papers in the SAME DOMAIN with experimental guidance:
        
        Rules:
        - Term 1: Primary domain-specific technique or model (e.g., "YOLO", "transformer", "CNN", "LSTM")
        - Term 2: Specific task type (e.g., "object detection", "classification", "segmentation")  
        - Term 3: Experimental aspect (e.g., "ablation study", "evaluation", "comparison", "benchmark")
        - Term 4: Domain/application context (e.g., "autonomous driving", "medical imaging", "NLP")
        
        Examples:
        - For computer vision: "YOLO/object detection/ablation study/autonomous driving"
        - For NLP: "transformer/text classification/evaluation/sentiment analysis"
        - For medical AI: "CNN/medical imaging/comparison/radiology"
        
        SPECIAL CASE FOR SELF-SUPERVISED LEARNING: If the research direction involves self-supervised learning, use terms like:
        - "SimCLR/contrastive learning/CIFAR-10/pretraining"
        - "BYOL/self-supervised learning/image classification/fine-tuning"
        - "MAE/masked autoencoder/CIFAR-10/generalization"
        - "MoCo/momentum contrast/CIFAR-10/representation learning"
        
        Focus on finding papers that will have similar experimental setups and methodologies, NOT generic methodology papers.
        Return ONLY the 4-term query string (no explanation).
        """
    
        # Use higher temperature for retries to get different results
        temperature = 0.3 if search_iteration > 0 else 0.1
        
        response = client.chat.completions.create(
            model=model,
            temperature=temperature,
            messages=[
                {"role": "system", "content": "Generate focused domain-specific ArXiv search queries. For retries, ensure the new query is significantly different from previous attempts."},
                {"role": "user", "content": query_prompt}
            ]
        )

        search_query = response.choices[0].message.content.strip()
        
        # Clean the query (remove quotes, extra spaces)
        search_query = search_query.replace('"', '').replace("'", "").strip()
        
        # Validate that retry query is actually different
        if search_iteration > 0 and previous_query and search_query == previous_query:
            _write_stream("Generated query is identical to previous - creating fallback different query...")
            
            # Create a guaranteed different query
            base_terms = previous_query.split('/') if '/' in previous_query else ["machine learning", "experimental", "evaluation", "methodology"]
            
            # Alternative search strategies for different domains
            if "computer vision" in primary_domain.lower() or "cv" in primary_domain.lower():
                alternative_queries = [
                    "ResNet/image classification/systematic evaluation/ImageNet",
                    "convolutional neural network/experimental analysis/performance comparison/computer vision",
                    "deep learning/ablation study/benchmark evaluation/visual recognition"
                ]
            elif "nlp" in primary_domain.lower() or "text" in primary_domain.lower():
                alternative_queries = [
                    "BERT/text classification/experimental evaluation/NLP",
                    "transformer/language model/systematic comparison/text analysis",
                    "neural language processing/empirical study/benchmark/natural language"
                ]
            else:
                alternative_queries = [
                    "neural network/experimental methodology/empirical evaluation/machine learning",
                    "deep learning/systematic study/performance analysis/artificial intelligence",
                    "machine learning/experimental validation/comparative study/methodology"
                ]
            
            # Pick different query based on iteration
            search_query = alternative_queries[search_iteration % len(alternative_queries)]
        
        # Ensure it has the right format (4 terms separated by /)
        if search_query.count('/') != 3:
            # Fallback: create domain-specific query from extracted info
            term1 = task_type or primary_domain
            term2 = "experimental" if not task_type else task_type
            term3 = "evaluation"
            term4 = application_area or primary_domain
            search_query = f"{term1}/{term2}/{term3}/{term4}"
        
        _write_stream(f"Generated {'refined' if search_iteration > 0 else 'initial'} search query: {search_query}")
        if search_iteration > 0 and previous_query:
            _write_stream(f"Previous query was: {previous_query}")
        
        return {
            **state,
            "experiment_search_query": search_query,
            "experiment_search_domain": primary_domain,
            "experiment_search_task": task_type,
            "current_step": "search_query_generated"
        }
        
    except Exception as e:
        print(f"Error generating experiment search query: {str(e)}")
        # Enhanced fallback for retries
        prompt_lower = original_prompt.lower()
        direction_lower = direction_text.lower()
        
        if search_iteration > 0:
            fallback_queries = [
                "experimental methodology/systematic evaluation/empirical study/research methodology",
                "performance analysis/comparative study/experimental validation/machine learning",
                "ablation study/experimental design/empirical analysis/systematic comparison"
            ]
            fallback_query = fallback_queries[search_iteration % len(fallback_queries)]
        elif "self-supervised" in direction_lower or "ssl" in direction_lower:
            fallback_query = "SimCLR/contrastive learning/CIFAR-10/pretraining"
        elif "object detection" in prompt_lower or "detection" in prompt_lower:
            fallback_query = "object detection/evaluation/experimental/computer vision"
        elif "classification" in prompt_lower:
            fallback_query = "classification/experimental/evaluation/machine learning"
        elif "segmentation" in prompt_lower:
            fallback_query = "segmentation/experimental/evaluation/computer vision"
        elif "nlp" in prompt_lower or "text" in prompt_lower:
            fallback_query = "text classification/experimental/evaluation/NLP"
        else:
            fallback_query = "machine learning/experimental/evaluation/methodology"
        
        return {
            **state,
            "experiment_search_query": fallback_query,
            "experiment_search_domain": "machine learning",
            "errors": state.get("errors", []) + [f"Search query generation error: {str(e)}"],
            "current_step": "search_query_error"
        }

async def _search_experiment_papers_node(state: ExperimentSuggestionState) -> ExperimentSuggestionState:
    """Search ArXiv for experimental methodology papers using optimized workflow."""
    search_iteration = state.get("experiment_search_iteration", 0)
    validation_results = state.get("experiment_paper_validation_results", {})
    is_backup_search = validation_results.get("decision") == "search_backup"
    
    if search_iteration == 0:
        _write_stream("Experiment Papers Search: Searching ArXiv for experimental guidance...")
    elif is_backup_search:
        _write_stream("Experiment Search (Backup): Searching for additional experimental papers...")
    else:
        _write_stream(f"Experiment Search (New Search {search_iteration + 1}): Searching with refined query...")
        
    state["current_step"] = "search_experiment_papers"
    
    # Import required modules for ArXiv search
    import urllib.request as libreq
    import xml.etree.ElementTree as ET
    from utils.arxiv import format_search_string
    from concurrent.futures import ThreadPoolExecutor, as_completed
    
    # Initialize variables
    papers = []
    total_results = 0
    formatted_query = ""
    
    # For backup searches, preserve existing papers
    existing_papers = []
    if is_backup_search and state.get("experiment_papers"):
        existing_papers = state["experiment_papers"]
        _write_stream(f"Preserving {len(existing_papers)} papers from previous search")
    
    # Safety check: After 3 iterations, force continue to avoid infinite loops
    if search_iteration >= 3:
        _write_stream(f"Maximum search iterations reached ({search_iteration}). Proceeding with existing papers...")
        state["experiment_papers"] = existing_papers if existing_papers else []
        state["experiment_search_completed"] = True
        return state
    
    try:
        search_query = state.get("experiment_search_query", "experimental methodology")
        research_direction = state.get("research_direction", {})
        original_prompt = state.get("original_prompt", "")
        
        # Use ArXiv processor for paper processing
        arxiv_processor = state["arxiv_processor"]
        if not arxiv_processor:
            raise Exception("ArXiv processor not available")
        
        # Determine search parameters based on search type and iteration
        if search_iteration == 0:
            # Initial search: get 100 papers for ranking
            max_results = 100
            start_offset = 0
        elif is_backup_search:
            # Backup search: get additional papers with offset
            existing_count = len(existing_papers) if existing_papers else 0
            start_offset = max(100, existing_count)
            max_results = 100
        else:
            # New search with different query: get 100 fresh papers
            max_results = 100
            start_offset = 0
        
        
        
        # Format search query and build URL
        formatted_query = format_search_string(search_query)
        url = f"http://export.arxiv.org/api/query?search_query={formatted_query}&start={start_offset}&max_results={max_results}"
        
        #_write_stream(f"Formatted query: {formatted_query}")
        _write_stream(f"Full search URL: {url}")
        
        # Fetch and parse ArXiv results
        with libreq.urlopen(url) as response:
            xml_data = response.read()
        
        root = ET.fromstring(xml_data)
        ns = {
            'atom': 'http://www.w3.org/2005/Atom',
            'arxiv': 'http://arxiv.org/schemas/atom',
            'opensearch': 'http://a9.com/-/spec/opensearch/1.1/'
        }
        
        # Get total results
        total_results_elem = root.find('opensearch:totalResults', ns)
        total_results = int(total_results_elem.text) if total_results_elem is not None else 0
        
        _write_stream(f"Total papers found: {total_results}")
        
        if total_results == 0:
            print("⚠️ No papers found for experiment search query")
            return {
                **state,
                "experiment_papers": existing_papers,
                "current_step": "no_papers_found"
            }
        
        # Extract paper entries
        entries = root.findall('atom:entry', ns)
        if len(entries) == 0:
            entries = root.findall('.//entry')  # Fallback without namespace
        
        _write_stream(f"Processing {len(entries)} paper entries...")
        
        # Stage 1: Extract basic info (title, abstract, metadata) without downloading PDFs
        _write_stream(f"Stage 1: Extracting basic info for {len(entries)} experimental papers...")
        papers = []
        for i, entry in enumerate(entries, 1):
            try:
                paper_info = arxiv_processor.extract_basic_paper_info(entry, ns, i)
                papers.append(paper_info)
               # print(f"✅ Basic info extracted for paper #{i}: {paper_info['title'][:50]}...")
            except Exception as e:
                print(f"⚠️ Error processing paper entry {i}: {e}")
                continue
        
        # Stage 2: Rank papers by relevance using enhanced analysis context
        _write_stream(f"Ranking experimental papers by relevance using extracted analysis...")
        
        # Create enhanced ranking context from the analysis findings
        # Note: These are utility functions that should be called as standalone functions
        ranking_context = create_experiment_ranking_context_from_analysis(state)
        #print(f"📊 Using enhanced context for ranking: {ranking_context[:100]}...")
        
        # Create custom prompt for experimental ranking
        custom_prompt = create_custom_ranking_prompt("experimental")
        
        papers = await arxiv_processor.rank_papers_by_relevance(papers, ranking_context, custom_prompt)
        
        # Stage 3: Download full content for top 5 papers only
        top_papers = papers[:5]  # Get top 5 papers
        
        _write_stream(f"Downloading full PDF content for top {len(top_papers)} experimental papers...")

        with ThreadPoolExecutor(max_workers=5) as executor:  # Limit concurrent downloads
            # Submit download tasks for top papers only
            future_to_paper = {
                executor.submit(arxiv_processor.download_paper_content, paper): paper 
                for paper in top_papers
            }
            
            # Collect results as they complete
            for future in as_completed(future_to_paper):
                updated_paper = future.result()
                # Update the paper in the original list
                for i, paper in enumerate(papers):
                    if paper['id'] == updated_paper['id']:
                        papers[i] = updated_paper
                        break
        
        _write_stream(f"PDF download stage completed for experimental papers.")
        ''''
        # Print ranked results
        print("\n" + "=" * 80)
        print("📋 RANKED EXPERIMENTAL PAPERS (by relevance):")
        print("=" * 80)
        
        for i, paper in enumerate(papers[:5], 1):  # Show top 5
            relevance_score = paper.get('relevance_score', 0)
            has_content = paper.get('pdf_downloaded', False)
            content_status = "📄 FULL CONTENT" if has_content else "📝 TITLE+ABSTRACT"
            
            print(f"\n📄 EXPERIMENTAL PAPER #{i} ({content_status}) - Relevance: {relevance_score:.1f}/10.0")
            print("-" * 60)
            print(f"Title: {paper['title']}")
            print(f"ID: {paper['id']}")
            print(f"Published: {paper['published']}")
            print(f"URL: {paper['url']}")
            
            if paper.get('summary'):
                print(f"Summary: {paper['summary'][:300]}...")
            if has_content and paper.get('content'):
                print(f"Content Snippet: {paper['content'][:500]}...")
            
            print("-" * 60)
        '''
        # Combine with existing papers if this is a backup search
        final_papers = papers
        if is_backup_search and existing_papers:
            # Merge papers, avoiding duplicates
            existing_ids = {p['id'] for p in existing_papers}
            new_papers = [p for p in papers if p['id'] not in existing_ids]
            final_papers = existing_papers + new_papers
            
            # Sort by relevance score and keep only top 5
            final_papers.sort(key=lambda x: x.get('relevance_score', 0), reverse=True)
            final_papers = final_papers[:7]
        else:
            # For non-backup searches, also limit to top 5
            final_papers = papers[:5]
        return {
            **state,
            "experiment_papers": final_papers,
            "experiment_search_iteration": search_iteration + 1,
            "current_step": "papers_downloaded"
        }
        
    except Exception as e:
        print(f"❌ Error in experiment papers search: {str(e)}")
        return {
            **state,
            "experiment_papers": [],
            "errors": state.get("errors", []) + [f"Experiment papers search error: {str(e)}"],
            "current_step": "search_error"
        }

def _validate_experiment_papers_node(state: ExperimentSuggestionState) -> ExperimentSuggestionState:
    """Node to validate if retrieved experiment papers can answer the user's query and decide next steps."""
    model= state["model"]
    client = state["client"]
    _write_stream("Validating experiment paper relevance and determining next steps.")
    state["current_step"] = "validate_experiment_papers"
    
    # Early bypass: If max iterations reached, skip validation and proceed directly
    search_iteration = state.get("experiment_search_iteration", 0)
    if search_iteration >= 3:
        _write_stream(f"⚠️ Maximum iterations ({search_iteration}) reached. Skipping validation and proceeding to experiment generation...")
        
        # CRITICAL FIX: Transfer papers to the correct state keys for new clean architecture
        papers = state.get("experiment_papers", [])
        #print(f"📚 Transferring {len(papers)} papers to validated_experiment_papers for clean architecture")
        
        state["experiment_paper_validation_decision"] = "PROCEED_DIRECT"
        state["validated_experiment_papers"] = papers  # NEW: Transfer papers to the key the clean architecture expects
        state["experiment_paper_validation_results"] = {
            "validation_result": "SKIP",
            "decision": "PROCEED_DIRECT", 
            "papers_count": len(papers),
            "reason": "max_iterations_reached"
        }
        state["next_node"] = "distill_paper_methodologies"  # Route to distillation step first
        return state
    
    try:
        papers = state.get("experiment_papers", [])
        user_query = state["original_prompt"]
        research_direction = state.get("research_direction", {})
        
        # Prepare paper summaries for validation with methodology checks
        papers_summary = ""
        full_content_papers = [p for p in papers if p.get('pdf_downloaded', False)]
        
        # Include information about all papers with enhanced methodology analysis
        methodology_count = 0
        experiment_count = 0
        
        for i, paper in enumerate(papers[:10], 1):  # Top 10 papers
            clean_title = _clean_text_for_utf8(paper.get('title', 'Unknown Title'))
            clean_abstract = _clean_text_for_utf8(paper.get('summary', 'No abstract available'))
            full_content = _clean_text_for_utf8(paper.get('content', ''))
            relevance_score = paper.get('relevance_score', 0)
            has_content = paper.get('pdf_downloaded', False)
            content_status = "FULL CONTENT" if has_content else "TITLE+ABSTRACT"
            
            # Check for methodology and experiments sections
            has_methodology = False
            has_experiments = False
            
            if has_content and full_content:
                content_lower = full_content.lower()
                # Look for methodology indicators
                methodology_keywords = ['methodology', 'method', 'approach', 'algorithm', 'procedure', 'framework', 'implementation']
                has_methodology = any(keyword in content_lower for keyword in methodology_keywords)
                
                # Look for experiment indicators
                experiment_keywords = ['experiment', 'evaluation', 'result', 'performance', 'benchmark', 'dataset', 'accuracy', 'precision', 'recall']
                has_experiments = any(keyword in content_lower for keyword in experiment_keywords)
                
                if has_methodology:
                    methodology_count += 1
                if has_experiments:
                    experiment_count += 1
            
            methodology_status = "✅ METHODOLOGY" if has_methodology else "❌ NO METHODOLOGY"
            experiment_status = "✅ EXPERIMENTS" if has_experiments else "❌ NO EXPERIMENTS"
            
            papers_summary += f"""
                Paper {i} [{content_status}] - Relevance: {relevance_score:.1f}/10.0:
                Title: {clean_title}
                Abstract: {clean_abstract}
                content_snippet: {full_content[:4000]}...
                Content Analysis: {methodology_status} | {experiment_status}
                ---
            """
        
        # Extract research direction context
        selected_direction = research_direction.get("selected_direction", {})
        direction_text = selected_direction.get("direction", "General experimental guidance")
        key_questions = selected_direction.get("key_questions", [])
        
        # Create enhanced validation prompt with HYPER-STRICT requirements and clear JSON format
        validation_prompt = f"""
            You are a HYPER-STRICT research analyst. Only papers with DETAILED METHODOLOGY and CONCRETE EXPERIMENTS should pass validation.

            USER'S QUERY: {_clean_text_for_utf8(user_query)}
            RESEARCH DIRECTION: {direction_text}
            KEY QUESTIONS: {', '.join(key_questions[:3]) if key_questions else 'General experimental guidance'}
            CURRENT SEARCH ITERATION: {search_iteration + 1}

            RETRIEVED PAPERS:
            {papers_summary}

            SEARCH STATISTICS:
            - Total papers found: {len(papers)}
            - Papers with full content: {len(full_content_papers)}
            - Papers with methodology sections: {methodology_count}
            - Papers with experiment sections: {experiment_count}
            - Average relevance score: {sum(p.get('relevance_score', 0) for p in papers) / len(papers) if papers else 0:.2f}/10.0

            HYPER-STRICT REQUIREMENTS FOR EXPERIMENTAL GUIDANCE PAPERS:
            1. **METHODOLOGY REQUIREMENT**: Papers MUST contain detailed experimental methodologies, algorithms, or implementation details
            2. **EXPERIMENTS REQUIREMENT**: Papers MUST contain actual experimental results, evaluations, or empirical validation
            3. **RELEVANCE REQUIREMENT**: Papers MUST be directly relevant to the research direction (≥8.0/10.0 for "continue")
            4. **COMPLETENESS REQUIREMENT**: Papers MUST provide actionable experimental guidance, not just theoretical discussions
            5. **TECHNICAL DEPTH REQUIREMENT**: Papers MUST include specific experimental procedures, datasets, metrics, or protocols

            STRICT DECISION CRITERIA:
            - "continue": ≥5 papers with BOTH methodology AND experiments, avg relevance ≥8.0, comprehensive experimental coverage
            - "search_backup": 3-4 papers with methodology/experiments, avg relevance ≥7.0, partial experimental coverage
            - "search_new": <3 papers with methodology/experiments, avg relevance <7.0, insufficient experimental guidance

            WARNING: Be EXTREMELY STRICT. Only "continue" if papers provide CONCRETE, ACTIONABLE experimental methodologies.

            REQUIRED JSON FORMAT (return ONLY this JSON, no other text):
            {{
                "relevance_assessment": "excellent" | "good" | "fair" | "poor",
                "methodology_coverage": "comprehensive" | "partial" | "insufficient",
                "experiment_coverage": "comprehensive" | "partial" | "insufficient", 
                "actionable_guidance": "high" | "medium" | "low",
                "technical_depth": "detailed" | "moderate" | "superficial",
                "decision": "continue" | "search_backup" | "search_new",
                "confidence": 0.95,
                "reasoning": "Brief explanation focusing on methodology and experimental content quality",
                "missing_aspects": ["aspect1", "aspect2", "aspect3"],
                "methodology_gaps": ["gap1", "gap2", "gap3"],
                "search_guidance": {{
                    "new_search_terms": ["term1", "term2", "term3"],
                    "focus_areas": ["area1", "area2", "area3"],
                    "avoid_terms": ["avoid1", "avoid2"]
                }}
            }}

            BE RUTHLESS: If papers lack concrete experimental methodologies or detailed experimental procedures, choose "search_new".
            Return only valid JSON with all required fields.
        """

        # Call LLM for validation
        response = client.chat.completions.create(
            model=model,
            temperature=0,
            messages=[
                {"content": validation_prompt, "role": "user"}
            ]
        )
        
        validation_response = response.choices[0].message.content.strip()
        
        # Parse validation response
        try:
            # Remove any markdown formatting
            if validation_response.startswith("```json"):
                validation_response = validation_response[7:]
            if validation_response.endswith("```"):
                validation_response = validation_response[:-3]
            validation_response = validation_response.strip()
            
            validation_data = json.loads(validation_response)
            
            # Store validation results in state - use unique key to avoid conflicts
            state["experiment_paper_validation_results"] = {
                "validation_successful": True,
                "validation_data": validation_data,
                "decision": validation_data.get("decision", "continue"),
                "reasoning": validation_data.get("reasoning", "No reasoning provided"),
                "missing_aspects": validation_data.get("missing_aspects", []),
                "search_guidance": validation_data.get("search_guidance", {}),
                "iteration": search_iteration + 1
            }
            
            # ALSO store decision in a separate key to avoid conflicts with other workflows
            state["experiment_paper_validation_decision"] = validation_data.get("decision", "continue")
            
        
            # Safe extraction with defaults and formatting
            relevance = str(validation_data.get('relevance_assessment', 'Unknown')).title()
            methodology_coverage = str(validation_data.get('methodology_coverage', 'Unknown')).title()
            experiment_coverage = str(validation_data.get('experiment_coverage', 'Unknown')).title()
            actionable_guidance = str(validation_data.get('actionable_guidance', 'Unknown')).title()
            technical_depth = str(validation_data.get('technical_depth', 'Unknown')).title()
            decision = str(validation_data.get('decision', 'continue')).upper()
            confidence = float(validation_data.get('confidence', 0))
            reasoning = str(validation_data.get('reasoning', 'No reasoning provided'))
            '''
            print(f"🎯 Relevance Assessment: {relevance}")
            print(f"🔬 Methodology Coverage: {methodology_coverage}")
            print(f"🧪 Experiment Coverage: {experiment_coverage}")
            print(f"📋 Actionable Guidance: {actionable_guidance}")
            print(f"⚙️ Technical Depth: {technical_depth}")
            print(f"🚀 Decision: {decision}")
            print(f"🎲 Confidence: {confidence:.2f}")
            print(f"💭 Reasoning: {reasoning}")
            '''
            # Handle missing aspects with proper formatting
            missing_aspects = validation_data.get('missing_aspects', [])
            '''
            if missing_aspects and isinstance(missing_aspects, list):
                print(f"🔍 Missing Experimental Aspects:")
                for i, aspect in enumerate(missing_aspects[:5], 1):  # Limit to 5 items
                    print(f"   {i}. {str(aspect)}")
            
            # Handle methodology gaps
            methodology_gaps = validation_data.get('methodology_gaps', [])
            if methodology_gaps and isinstance(methodology_gaps, list):
                print(f"🔧 Methodology Gaps:")
                for i, gap in enumerate(methodology_gaps[:3], 1):  # Limit to 3 items
                    print(f"   {i}. {str(gap)}")
            '''
            # Handle search guidance for non-continue decisions
            if decision != 'CONTINUE':
                search_guidance = validation_data.get('search_guidance', {})
                if isinstance(search_guidance, dict):
                    new_search_terms = search_guidance.get('new_search_terms', [])
                    focus_areas = search_guidance.get('focus_areas', [])
                    avoid_terms = search_guidance.get('avoid_terms', [])
                    
                    if new_search_terms and isinstance(new_search_terms, list):
                        _write_stream(f"Suggested Search Terms: {', '.join(str(term) for term in new_search_terms[:7])}")
                    if focus_areas and isinstance(focus_areas, list):
                        _write_stream(f"Focus Areas: {', '.join(str(area) for area in focus_areas[:5])}")
                    if avoid_terms and isinstance(avoid_terms, list):
                        _write_stream(f"Avoid Terms: {', '.join(str(term) for term in avoid_terms[:5])}")
            
   
            # CRITICAL FIX: Make the routing decision here instead of in separate function
            validation_decision = validation_data.get("decision", "continue").upper()
            
            # Map validation decision to next node
            if validation_decision == "CONTINUE":
                next_node = "distill_paper_methodologies"  # Route to distillation step first
                _write_stream(f"Experiment papers are adequate. Continuing to experiment suggestions.")
                
                state["validated_experiment_papers"] = papers
                
                
            elif validation_decision == "SEARCH_BACKUP":
                next_node = "search_experiment_papers"
                _write_stream(f"Papers need backup. Searching for additional papers.")
            elif validation_decision == "SEARCH_NEW":
                next_node = "generate_experiment_search_query"
                _write_stream(f"Papers inadequate. Generating new search query.")
            else:
                # Default fallback
                next_node = "distill_paper_methodologies"  # Route to distillation step first
                _write_stream(f"Unknown validation decision '{validation_decision}'. Defaulting to continue.")
                
            
                state["validated_experiment_papers"] = papers
            
            # Increment search iteration counter
            state["experiment_search_iteration"] = search_iteration + 1
            state["next_node"] = next_node
            
            # Return state after successful validation
            return state
            
        except json.JSONDecodeError as e:
            error_msg = f"Failed to parse validation JSON: {e}"
            print(f"ERROR: {error_msg}")
            
            # Fallback decision based on paper quality
            avg_score = sum(p.get('relevance_score', 0) for p in papers) / len(papers) if papers else 0
            decision = "continue" if avg_score >= 6.0 else "search_backup"
            
            state["experiment_paper_validation_results"] = {
                "validation_successful": False,
                "error": error_msg,
                "decision": decision,
                "reasoning": f"Fallback decision based on average score: {avg_score:.2f}",
                "iteration": search_iteration + 1
            }
            
            # ALSO store decision in backup key for error cases
            state["experiment_paper_validation_decision"] = decision
            
            # Add routing for error case
            if decision == "continue":
                state["next_node"] = "distill_paper_methodologies"  # Route to distillation step first
                # CRITICAL FIX: Transfer papers for continue decision in JSON error case
               
                state["validated_experiment_papers"] = papers
            else:
                state["next_node"] = "search_experiment_papers"
            
            state["experiment_search_iteration"] = search_iteration + 1
            
            
    except Exception as e:
        error_msg = f"Validation failed: {str(e)}"
        print(f"ERROR: {error_msg}")
        
        # Default to continue on error
        state["experiment_paper_validation_results"] = {
            "validation_successful": False,
            "error": error_msg,
            "decision": "continue",
            "reasoning": "Error occurred, defaulting to continue",
            "iteration": state.get("experiment_search_iteration", 0) + 1
        }
        
        # ALSO store decision in backup key for error cases
        state["experiment_paper_validation_decision"] = "continue"
        
        # Add routing for error case  
        state["next_node"] = "distill_paper_methodologies"  # Route to distillation step first
        
        # CRITICAL FIX: Transfer papers for general error case too
        papers = state.get("experiment_papers", [])
       # print(f"📚 Transferring {len(papers)} papers to validated_experiment_papers (general error)")
        state["validated_experiment_papers"] = papers
        
        state["experiment_search_iteration"] = state.get("experiment_search_iteration", 0) + 1
    
    return state

async def _distill_paper_methodologies_node(state: ExperimentSuggestionState) -> ExperimentSuggestionState:
    """Distill and condense methodology and experimental information from validated papers."""
    _write_stream("Distilling methodology and experimental information from papers...")
    state["current_step"] = "distill_methodologies"

    client = state["client"]
    model = state["model"]
    validated_papers = state.get("validated_experiment_papers", [])

    if not validated_papers:
        print("⚠️ No validated papers to distill")
        state["distilled_methodologies"] = {}
        return state

    try:
        distilled_methodologies = {}

        for i, paper in enumerate(validated_papers[:5], 1):  # Process top 5 papers
            paper_title = paper.get('title', f'Paper {i}')
            paper_content = paper.get('content', '')

            if not paper_content:
                #print(f"⚠️ Skipping paper {i} - no content available")
                continue

            #print(f"🔬 Distilling paper {i}: {paper_title[:50]}...")

            # Create distillation prompt with clearer instructions for brevity
            distillation_prompt = f"""Extract key methodology from this research paper in exactly 800 characters or less.

FOCUS ON:
• Model architecture (specific networks, layers, components)
• Training setup (batch size, learning rate, optimizer, epochs)
• Dataset and preprocessing
• Key experimental details

FORMAT: Bullet points only. Be concise but specific.

PAPER: {paper_title}
CONTENT: {paper_content}
IMPORTANT:
KEEP RESPONSES TO A MAXIMUM OF 600 CHARACTERS.
Provide methodology summary in under 600 characters:"""

            try:
                # Remove max_tokens to avoid truncation issues with LiteLLM proxy
                response = client.chat.completions.create(
                    model=model,
                    temperature=0.1,  # Low temperature for factual extraction
                    messages=[{"content": distillation_prompt, "role": "user"}]
                )

                distilled_content = response.choices[0].message.content
                '''
                # Debug response info
                print(f"🔍 API response finish_reason: {response.choices[0].finish_reason}")
                print(f"🔍 Raw content type: {type(distilled_content)}")
                print(f"🔍 Raw content length: {len(distilled_content) if distilled_content else 0}")
                '''
                # Handle None content
                if distilled_content is None:
                    print(f"⚠️ API returned None content for paper {i}")
                    distilled_content = "API returned no content - processing failed"
                
                # Ensure content is within 1200 character limit
                if distilled_content and len(distilled_content) > 1200:
                    distilled_content = distilled_content[:1197] + "..."

                distilled_methodologies[f"paper_{i}"] = {
                    "title": paper_title,
                    "distilled_content": distilled_content or "No methodology information extracted",
                    "character_count": len(distilled_content) if distilled_content else 0
                }

                _write_stream(f"Distilled paper {i}: {len(distilled_content) if distilled_content else 0} characters")
                
            except Exception as api_error:
                print(f"API error for paper {i}: {api_error}")
                distilled_methodologies[f"paper_{i}"] = {
                    "title": paper_title,
                    "distilled_content": f"API error: {str(api_error)}",
                    "character_count": 0
                }

        state["distilled_methodologies"] = distilled_methodologies
        for i in range(len(validated_papers), 5):
            distilled_methodologies[f"paper_{i+1}"] = {
                "title": f"Paper {i+1}",
                "distilled_content": "No paper available",
                "character_count": 0
            }

        total_chars = sum(info['character_count'] for info in distilled_methodologies.values())
        successful_distillations = sum(1 for info in distilled_methodologies.values() if info['character_count'] > 0)
        
        #print(f"📚 Successfully distilled methodologies from {successful_distillations}/{len(distilled_methodologies)} papers")
        _write_stream(f"Total distilled content: {total_chars} characters")

    except Exception as e:
        error_msg = f"Methodology distillation failed: {str(e)}"
        print(f"ERROR: {error_msg}")
        state["errors"].append(error_msg)
        state["distilled_methodologies"] = {}
        # Continue even on error - workflow will proceed to experiment generation

    return state

async def _suggest_experiments_tree_2_node(state: ExperimentSuggestionState) -> ExperimentSuggestionState:
    """Generate comprehensive experiment suggestions."""
    # Extract dependencies from state
    client = state["client"]
    model = state["model"]

    
    _write_stream("Generating experiment suggestions.")
    state["current_step"] = "suggest_experiments"
    
    try:
        research_direction = state.get("research_direction", {})
        findings_analysis = state.get("findings_analysis", {})
        validated_papers = state.get("validated_experiment_papers", [])
        original_prompt = state["original_prompt"]

        # Get validation feedback from previous iteration
        validation_feedback = state.get("validation_feedback", "")
        current_iteration = state.get("current_experiment_iteration", 0)

        # Store current experiment suggestions as previous before generating new ones
        current_suggestions = state.get("experiment_suggestions", "")
        if current_suggestions and current_iteration > 0:
            state["previous_experiment_suggestions"] = current_suggestions
           # print(f"📝 Stored previous experiment suggestions ({len(current_suggestions)} chars) for improvement guidance")
        
        # Get accumulated past mistakes for learning
        past_mistakes = state.get("past_experiment_mistakes", [])
        
        # Get previous experiment suggestions for direct improvement
        previous_experiment_suggestions = state.get("previous_experiment_suggestions", "")
        
        # Prepare context from distilled methodologies instead of raw papers
        distilled_methodologies = state.get("distilled_methodologies", {})
        papers_context = ""
        
        # Check if we have meaningful distilled content (not just empty entries)
        has_meaningful_distillation = any(
            info.get('distilled_content', '') and len(info.get('distilled_content', '')) > 10
            for info in distilled_methodologies.values()
        )
        
        if has_meaningful_distillation:
            papers_context = "Distilled Methodology and Experimental Information from Research Papers:\n\n"
            for paper_key, paper_info in distilled_methodologies.items():
                papers_context += f"Paper: {paper_info['title']}\n"
                papers_context += f"Distilled Content ({paper_info['character_count']} chars):\n"
                papers_context += f"{paper_info['distilled_content']}\n\n"
        elif validated_papers:
            # Fallback to raw papers if distillation failed or produced no content
            _write_stream("Using raw paper content as fallback (distillation produced no meaningful content)")
            papers_context = "Relevant Research Papers:\n"
            for i, paper in enumerate(validated_papers[:5], 1):
                papers_context += f"{i}. {paper.get('title', 'Unknown')}\n"
                papers_context += f"Content: {paper.get('content', 'No Content')[:2000]}...\n\n"
        else:
            papers_context = "No paper context available for experiment generation."
        
        # Add improvement guidance for iterations with past mistakes history
        improvement_guidance = ""
        if current_iteration > 0:
            improvement_parts = []
            
            if validation_feedback:
                improvement_parts.append(f"""
        CRITICAL IMPROVEMENT REQUIREMENTS (Iteration {current_iteration + 1}):
        Previous validation feedback: {validation_feedback}""")
            
            # Include detailed history of past mistakes to prevent repetition
            if past_mistakes:
                improvement_parts.append(f"""
        
        PAST MISTAKES HISTORY (LEARN FROM THESE - DO NOT REPEAT):
        You have failed validation {len(past_mistakes)} time(s) before. Study these mistakes carefully:""")
                
                for i, mistake in enumerate(past_mistakes, 1):
                    improvement_parts.append(f"""
        Iteration {mistake['iteration']} Failure (Score: {mistake['validation_score']:.2f}):
        - Critical Issues: {'; '.join(mistake['critical_issues'][:3]) if mistake['critical_issues'] else 'None'}
        - Direction Problems: {'; '.join(mistake['direction_misalignment'][:2]) if mistake['direction_misalignment'] else 'None'}
        - Novelty Issues: {'; '.join(mistake['novelty_concerns'][:2]) if mistake['novelty_concerns'] else 'None'}
        - Required Fixes: {'; '.join(mistake['improvement_recommendations'][:3]) if mistake['improvement_recommendations'] else 'None'}""")
            
            # Include the most recent experiment suggestions for direct improvement
            if previous_experiment_suggestions:
                improvement_parts.append(f"""
        
        PREVIOUS EXPERIMENT SUGGESTIONS (IMPROVE UPON THESE):
        Here are your most recent experiment suggestions that failed validation. Study them carefully and make targeted improvements:
        
        {previous_experiment_suggestions}
        
        Key areas to improve based on previous attempt:
        - Address any missing or incomplete sections
        - Enhance technical depth and specificity
        - Improve paper integration and citations
        - Strengthen methodology and evaluation procedures
        - Fix any structural or content issues identified in validation""")
            
            improvement_parts.append("""
        
        To pass strict validation, ensure:
        1. Design 2 comprehensive experiments.
        2. Include ALL required sections: **Objective**, **Hypothesis**, **Methodology**, **Expected Outcomes**, **Success Metrics**
        3. Integrate insights from at least 3 research papers with specific citations
        4. Provide detailed technical depth (datasets, models, training procedures, evaluation metrics)
        5. Include implementation timeline, resource requirements, and risk assessment
        6. Add statistical analysis and validation procedures
        7. Address ALL previous validation failures - do not repeat past mistakes""")
            
            improvement_guidance = "".join(improvement_parts)
        
        # Create experiment suggestion prompt with LITERATURE-GROUNDED but FLEXIBLE approach
        first_experiment_prompt = f"""
        You are an expert experimental researcher designing novel experiments grounded in existing literature. You can use techniques, models, and datasets from the provided papers, and combine them in novel ways to address the research questions.

        FLEXIBLE LITERATURE GROUNDING:
        - Use models, datasets, and techniques explicitly mentioned in the provided papers
        - Combine different approaches from the literature in novel ways
        - Reference specific papers for all experimental components
        - Create experiments that build upon and extend the existing work
        - If needed, use well-established techniques that complement the literature

        CRITICAL REQUIREMENTS:
        - Every experiment MUST reference at least 2 different papers from the provided literature
        - Models and datasets MUST be cited from the papers (e.g., "ResNet-50 (He et al., 2016)")
        - Experiments should address the key research questions directly
        - Include all required sections: Objective, Hypothesis, Methodology, Expected Outcomes, Success Metrics

        Original Research Context:
        {original_prompt}

        Research Direction:
        {json.dumps(research_direction, indent=2)}

        Findings Analysis:
        {json.dumps(findings_analysis, indent=2)}

        LITERATURE CONTEXT (MANDATORY REFERENCE MATERIAL):
        {papers_context}

        {improvement_guidance}

        EXPERIMENT DESIGN REQUIREMENTS:

        1. **LITERATURE GROUNDING**: Every experiment component MUST reference specific content from the provided papers
        2. **MODEL CITATIONS**: Every model mentioned MUST include original paper citation
        3. **DATASET CITATIONS**: Every dataset MUST be real with proper academic citation
        4. **TECHNIQUE VALIDATION**: Only use techniques explicitly described in the literature
        5. **PAPER INTEGRATION**: Reference at least 2 different papers for this experiment

        Generate ONE detailed experiment in the following format:

        ### Experiment 1: [Experiment Title - Reference specific paper technique]
        **Objective**: Clear statement grounded in literature findings
        **Hypothesis**: Testable hypothesis based on literature insights
        **Methodology**:
        - Use ONLY techniques from provided papers
        - Cite specific papers for each methodological choice
        - Reference exact models/datasets from literature
        **Expected Outcomes**: Based on literature expectations
        **Success Metrics**: Metrics mentioned in the papers
        **Resources Needed**: Based on literature resource requirements
        **Risk Assessment**: Risks identified in the literature
        **Literature References**: Specific papers supporting this experiment

        CRITICAL: Your experiment MUST be 100% grounded in the provided literature. If you cannot design an experiment using ONLY the content from these papers, state this explicitly rather than introducing external knowledge.
        """

        _write_stream("Generating first experiment...")
        response1 = client.chat.completions.create(
            model=model,
            temperature=0.3,
           # max_tokens=4000,
            messages=[{"content": first_experiment_prompt, "role": "user"}]
        )

        first_experiment = response1.choices[0].message.content

        # Ensure we have a valid string response
        if first_experiment is None:
            raise ValueError("LLM returned None response for first experiment")

        first_experiment = str(first_experiment)
        _write_stream("First experiment generated successfully")

        # Create experiment suggestion prompt for SECOND EXPERIMENT - ensuring novelty
        second_experiment_prompt = f"""
        You are an expert experimental researcher designing a SECOND experiment that complements the first but uses different approaches from the literature. You can combine techniques from the provided papers in novel ways to create a distinct experimental approach.

        FLEXIBLE LITERATURE GROUNDING FOR SECOND EXPERIMENT:
        - Use different models, datasets, or techniques than the first experiment
        - Combine approaches from the literature in novel ways
        - Reference specific papers for all experimental components
        - Create experiments that explore different aspects of the research questions
        - Build upon the literature while exploring alternative approaches

        CRITICAL REQUIREMENTS FOR SECOND EXPERIMENT:
        - MUST be fundamentally different from the first experiment in approach or methodology
        - Every experiment MUST reference at least 2 different papers from the provided literature
        - Models and datasets MUST be cited from the papers
        - Should address different aspects of the key research questions
        - Include all required sections: Objective, Hypothesis, Methodology, Expected Outcomes, Success Metrics

        FIRST EXPERIMENT (DO NOT REPEAT OR BE SIMILAR TO THIS):
        {first_experiment}

        Original Research Context:
        {original_prompt}

        Research Direction:
        {json.dumps(research_direction, indent=2)}

        Findings Analysis:
        {json.dumps(findings_analysis, indent=2)}

        LITERATURE CONTEXT (MANDATORY REFERENCE MATERIAL):
        {papers_context}

        {improvement_guidance}

        EXPERIMENT DESIGN REQUIREMENTS FOR SECOND EXPERIMENT:

        1. **LITERATURE GROUNDING**: Every experiment component MUST reference specific content from the provided papers
        2. **MODEL CITATIONS**: Every model mentioned MUST include original paper citation
        3. **DATASET CITATIONS**: Every dataset MUST be real with proper academic citation
        4. **TECHNIQUE VALIDATION**: Only use techniques explicitly described in the literature
        5. **PAPER INTEGRATION**: Reference at least 2 different papers for this experiment
        6. **NOVELTY REQUIREMENT**: This experiment MUST be fundamentally different from the first experiment in approach, methodology, models, or research focus
        7. **DIFFERENT LITERATURE BASIS**: Use different papers or different aspects of the same papers than the first experiment

        Generate ONE detailed experiment that explores a completely different angle or approach:

        ### Experiment 2: [Experiment Title - Reference specific paper technique - DIFFERENT from Experiment 1]
        **Objective**: Clear statement grounded in literature findings (different focus from Experiment 1)
        **Hypothesis**: Testable hypothesis based on literature insights (different from Experiment 1)
        **Methodology**:
        - Use ONLY techniques from provided papers (different techniques than Experiment 1)
        - Cite specific papers for each methodological choice (different papers/aspects than Experiment 1)
        - Reference exact models/datasets from literature (different from Experiment 1)
        **Expected Outcomes**: Based on literature expectations (different outcomes than Experiment 1)
        **Success Metrics**: Metrics mentioned in the papers (different metrics than Experiment 1)
        **Resources Needed**: Based on literature resource requirements
        **Risk Assessment**: Risks identified in the literature
        **Literature References**: Specific papers supporting this experiment (different papers than Experiment 1)

        CRITICAL REQUIREMENTS:
        - Your experiment MUST be 100% grounded in the provided literature
        - KEEP RESPONSE TO 500 WORDS MAXIMUM
        - Your experiment MUST be NOVEL and DIFFERENT from the first experiment
        - If you cannot design a sufficiently different experiment using ONLY the content from these papers, state this explicitly rather than creating a similar experiment
        """

        _write_stream("Generating second experiment (ensuring novelty)...")
        response2 = client.chat.completions.create(
            model=model,
            temperature=0.4,  # Slightly higher temperature for more creativity in second experiment
           # max_tokens=4000,
            messages=[{"content": second_experiment_prompt, "role": "user"}]
        )

        second_experiment = response2.choices[0].message.content

        # Ensure we have a valid string response
        if second_experiment is None:
            raise ValueError("LLM returned None response for second experiment")

        second_experiment = str(second_experiment)
        _write_stream("second experiment generated successfully")

        # Combine both experiments into final suggestions
        experiment_suggestions = f"""
# Experiment Design Recommendations

## Primary Experiments

{first_experiment}

{second_experiment}

## Quality Control Measures

- Statistical methods from the literature
- Reproducibility requirements mentioned in papers
- Validation procedures from literature

## Expected Timeline and Deliverables

- Timeline based on literature implementation times
- Deliverables aligned with literature outcomes
"""

        _write_stream("Both experiments combined successfully")
        
        state["experiment_suggestions"] = experiment_suggestions
        state["suggestion_source"] = "llm_generated"
        state["current_experiment_iteration"] = state.get("current_experiment_iteration", 0) + 1
        
        # Create summary with error handling
        try:
            total_experiments = len(re.findall(r'### Experiment \d+:', experiment_suggestions))
        except (TypeError, AttributeError) as e:
            print(f"⚠️ Error counting experiments: {e}")
            total_experiments = 0
        
        state["experiment_summary"] = {
            "total_experiments": total_experiments,
            "research_direction": research_direction.get("recommended_direction", {}).get("title", "Unknown"),
            "papers_used": len(validated_papers),
            "generation_successful": True,
            "iteration": state["current_experiment_iteration"]
        }
        
        _write_stream("Experiment suggestions generated successfully")
        _write_stream(f"Experiments designed: {state['experiment_summary']['total_experiments']}")
        
        # Set next node for workflow routing
        state["next_node"] = "validate_experiments_tree_2"
        
        # Add success message
        state["messages"].append(
            AIMessage(content=f"Generated two distinct experiment suggestions with {state['experiment_summary']['total_experiments']} detailed experiments, ensuring novelty between approaches.")
        )
        
    except Exception as e:
        error_msg = f"Experiment suggestion generation failed: {str(e)}"
        print(f"Error: {error_msg}")
        state["errors"].append(error_msg)
        
        # Create fallback suggestions
        state["experiment_suggestions"] = f"""
        # Experiment Design Recommendations

        ## Primary Experiment

        ### Experiment 1: Follow-up Validation Study
        **Objective**: Validate and extend the findings from the original research
        **Hypothesis**: The observed results can be reproduced and generalized
        **Methodology**: 
        - Replicate original experimental setup with variations
        - Expand dataset or parameter ranges
        - Apply statistical validation methods
        **Expected Outcomes**: Confirmation of original findings with extended insights
        **Success Metrics**: Reproducibility score > 0.8, statistical significance p < 0.05
        **Resources Needed**: Similar to original experiment setup
        **Risk Assessment**: Low risk, builds directly on existing work

        ## Implementation Roadmap

        **Phase 1**: Setup and replication (2 weeks)
        **Phase 2**: Extension experiments (2 weeks)  
        **Phase 3**: Analysis and validation (1 week)

        Note: This is a fallback experiment design. For detailed recommendations, please provide more specific experimental context.
        """
        
        state["experiment_summary"] = {
            "total_experiments": 1,
            "research_direction": "Fallback validation study",
            "papers_used": 0,
            "generation_successful": False,
            "iteration": 1
        }
        
        # Set next node even for fallback
        state["next_node"] = "validate_experiments_tree_2"
    
    return state


def _validate_experiments_tree_2_node(state: ExperimentSuggestionState) -> ExperimentSuggestionState:
    """Validate the generated experiment suggestions and decide whether to finalize or iterate."""
    _write_stream("Validating experiment suggestions.")
    state["current_step"] = "validate_experiments_tree_2"
    
    # Add infinite loop protection
    current_iteration = state.get("current_experiment_iteration", 0)
    if current_iteration >= 7:  # Higher safety limit
        _write_stream(f"Maximum experiment iterations reached ({current_iteration}). Forcing completion...")
        state["next_node"] = "END"
        state["experiments_validation_decision"] = "FORCE_PASS"
        return state
    experiment_suggestions = state.get("experiment_suggestions", "")
    
    if not experiment_suggestions:
        _write_stream("Experiment suggestions validation failed - no suggestions")
        state["next_node"] = "suggest_experiments_tree_2"
        state["validation_feedback"] = "No experiment suggestions generated"
        return state
    
    try:
        # Extract context for LLM validation
        experiment_summary = state.get("experiment_summary", {})
        total_experiments = experiment_summary.get("total_experiments", 0)
        papers_used = experiment_summary.get("papers_used", 0)
        
        # Get research direction and domain context for comprehensive validation
        research_direction = state.get("research_direction", {})
        findings_analysis = state.get("findings_analysis", {})
        validated_papers = state.get("validated_experiment_papers", [])
        original_prompt = state.get("original_prompt", "")

        # DEBUG: Check validated_papers state
       
        # Extract research direction details
        selected_direction = research_direction.get("selected_direction", {})
        direction_text = selected_direction.get("direction", "")
        key_questions = selected_direction.get("key_questions", [])
        direction_justification = selected_direction.get("justification", "")
        expected_impact = selected_direction.get("expected_impact", "")
        
        # Extract domain information
        domain_analysis = findings_analysis.get("domain_analysis", {})
        primary_domain = domain_analysis.get("primary_domain", "machine learning")
        task_type = domain_analysis.get("task_type", "")
        application_area = domain_analysis.get("application_area", "")
        model = state.get("model", "gemini/gemini-2.5-flash")
        model="gemini/gemini-2.5-flash-lite"
        client = state.get("client")
        
        # Build literature context from validated papers (use same format as original generation)
        literature_context = ""
        if validated_papers:
            literature_parts = []
            for i, paper in enumerate(validated_papers[:5], 1):
                title = paper.get("title", f"Paper {i}")
                abstract = paper.get("summary", paper.get("abstract", ""))
                content = paper.get("content", "")
                
                paper_content = f"**PAPER {i}: {title}**\n"
                
                if content:
                    paper_content += f"Key Content: {content}\n"
                
                literature_parts.append(paper_content)
            
            literature_context = "\n\n".join(literature_parts)
        
        # Get current iteration to match original generation prompt format
        current_iteration = state.get("current_experiment_iteration", 1)
        iteration_context = f"ITERATION {current_iteration}" if current_iteration > 1 else "INITIAL GENERATION"
        
        # SIMPLIFIED LLM-BASED VALIDATION
        validation_prompt = f"""
                You are an expert experimental methodology validator. Evaluate the proposed experiments for technical soundness, literature grounding, and research alignment.

                **RESEARCH CONTEXT:**
                - Direction: {direction_text}
                - Key Questions: {chr(10).join(f"• {q}" for q in key_questions[:3])}
                - Original Request: {original_prompt}

                **LITERATURE CONTEXT:**
                You have {len(validated_papers)} research papers that MUST be the foundation for all experiments:

                {literature_context}

                **PROPOSED EXPERIMENTS:**
                {experiment_suggestions}

                **VALIDATION REQUIREMENTS:**

                1. **LITERATURE GROUNDING (40% weight):**
                   - Every experiment component MUST reference specific content from the provided papers
                   - Use exact methodologies, datasets, and models mentioned in the literature
                   - Cite papers properly (e.g., "as described in Paper 1" or specific technique names)
                   - NO external techniques or models not mentioned in the papers

                2. **RESEARCH ALIGNMENT (35% weight):**
                   - Experiments MUST directly address the research direction: {direction_text}
                   - Each experiment should help answer the key questions
                   - Stay within the scope of transfer learning and fine-tuning strategies

                3. **TECHNICAL SOUNDNESS (25% weight):**
                   - Clear objectives, hypotheses, and methodologies
                   - Realistic resource requirements and timelines
                   - Proper evaluation metrics and success criteria
                   - Implementable experimental procedures

                **MANDATORY REQUIREMENTS:**
                - Every model/dataset MUST be cited from the provided papers
                - Experiments must include: Objective, Hypothesis, Methodology, Expected Outcomes, Success Metrics
                - At least 2 experiments must be proposed
                - Experiments should build upon the literature rather than require groundbreaking novelty

                **NOVELTY ASSESSMENT:**
                - For literature-grounded experiments, focus on practical combinations and extensions of documented methods
                - Value lies in systematic application and comparison of established approaches
                - Reasonable novelty comes from combining different techniques or applying them to new contexts within the literature

                Return your assessment in this exact JSON format:
                {{
                    "validation_result": "PASS" | "FAIL",
                    "overall_score": 0.0-1.0,
                    "detailed_scores": {{
                        "research_direction_alignment": 0.0-1.0,
                        "novelty_potential": 0.0-1.0,
                        "justification_quality": 0.0-1.0
                    }},
                    "critical_issues": ["list", "of", "critical", "problems"],
                    "direction_misalignment": ["ways", "experiments", "dont", "align"],
                    "novelty_concerns": ["lack", "of", "novelty", "issues"],
                    "improvement_recommendations": ["specific", "actionable", "suggestions"],
                    "decision_rationale": "Clear explanation of pass/fail decision"
                }}

                IMPORTANT: Respond with ONLY the JSON object, nothing else.
                """

        _write_stream("Starting LLM-based hyper-strict validation...")
        
       # print('VALIDAITON rompt: '+ validation_prompt)
        #print ("===="*30)
        
        # Initialize default values for error handling
        validation_result = "FAIL"
        overall_score = 0.0
        detailed_scores = {"research_direction_alignment": 0.0, "novelty_potential": 0.0, "justification_quality": 0.0}
        critical_issues = ["Validation initialization failed"]
        direction_misalignment = []
        novelty_concerns = []
        improvement_recommendations = ["Fix validation system", "Regenerate experiments"]
        decision_rationale = "Validation failed to initialize"
        validation_json = {}
        
        # SPECIAL HANDLING FOR NO PAPERS CASE
        if len(validated_papers) == 0:
            print(f"⚠️ No validated papers available for literature context. Using simplified validation criteria...")
            
            # Check if experiments have basic required sections
            experiment_content = experiment_suggestions.lower()
            required_sections = ["objective", "methodology", "resource", "timeline"]
            sections_found = sum(1 for section in required_sections if section in experiment_content)
            
            # Basic content validation
            has_experiments = len(re.findall(r'experiment \d+', experiment_content)) >= 2
            has_reasonable_length = len(experiment_suggestions) > 1000
            has_methodology = any(word in experiment_content for word in ['dataset', 'model', 'training', 'evaluation'])
            
            basic_score = 0.0
            if has_experiments:
                basic_score += 0.3
            if has_reasonable_length:
                basic_score += 0.2
            if has_methodology:
                basic_score += 0.3
            if sections_found >= 3:
                basic_score += 0.2
            
            print(f"📊 Basic validation score: {basic_score:.2f}/1.0")
            print(f"   - Has multiple experiments: {has_experiments}")
            print(f"   - Has reasonable length: {has_reasonable_length}")
            print(f"   - Has methodology: {has_methodology}")
            print(f"   - Required sections found: {sections_found}/4")
            
            if basic_score >= 0.7:  # Lower threshold when no papers available
                _write_stream("Experiment suggestions pass basic validation (no papers scenario)")
                overall_score = basic_score
                validation_result = "PASS"
                critical_issues = []
                improvement_recommendations = ["Add more baseline comparisons", "Include ablation studies", "Consider hyperparameter optimization"]
                decision_rationale = f"Basic validation passed with score {basic_score:.2f}/1.0 (no literature papers available)"
                validation_json = {
                    "validation_result": "PASS",
                    "overall_score": basic_score,
                    "detailed_scores": {"research_direction_alignment": basic_score, "novelty_potential": basic_score, "justification_quality": basic_score},
                    "critical_issues": [],
                    "direction_misalignment": [],
                    "novelty_concerns": [],
                    "improvement_recommendations": improvement_recommendations,
                    "decision_rationale": decision_rationale
                }
            else:
                _write_stream(f"Experiment suggestions failed basic validation - Score: {basic_score:.2f}/1.0")
                validation_result = "FAIL"
                overall_score = basic_score
                critical_issues = ["Insufficient experiment detail", "Missing required sections"]
                improvement_recommendations = ["Add more detailed methodology", "Include all required roadmap sections", "Expand experiment descriptions"]
                decision_rationale = f"Basic validation failed with score {basic_score:.2f}/1.0"
                validation_json = {
                    "validation_result": "FAIL",
                    "overall_score": basic_score,
                    "detailed_scores": {"research_direction_alignment": basic_score, "novelty_potential": basic_score, "justification_quality": basic_score},
                    "critical_issues": critical_issues,
                    "direction_misalignment": [],
                    "novelty_concerns": [],
                    "improvement_recommendations": improvement_recommendations,
                    "decision_rationale": decision_rationale
                }
        else:
            # Standard LLM validation when papers are available
            # Call LLM for validation (access client from state)
            try:
                client = state.get('client')
                if not client:
                    print("OpenAI client not found in state")
                    return state
                    
                response = client.chat.completions.create(
                    model=model,
                    temperature=0.1,
                    max_tokens=8000,
                    messages=[
                        {"role": "system", "content": f"You are a ruthlessly strict experimental methodology validator specializing in {primary_domain}. Respond with VALID JSON only - no markdown, no explanations, just the JSON object with all required fields."},
                        {"role": "user", "content": validation_prompt}
                    ]
                )
                
                validation_response = response.choices[0].message.content.strip()
                
                # Remove markdown formatting if present
                if validation_response.startswith("```json"):
                    validation_response = validation_response[7:]
                if validation_response.endswith("```"):
                    validation_response = validation_response[:-3]
                validation_response = validation_response.strip()
                
                # Parse validation JSON
                validation_json = json.loads(validation_response)
                
                # Extract results
                validation_result = validation_json.get("validation_result", "FAIL").upper()
                overall_score = float(validation_json.get("overall_score", 0.0))
                detailed_scores = validation_json.get("detailed_scores", detailed_scores)
                critical_issues = validation_json.get("critical_issues", [])
                direction_misalignment = validation_json.get("direction_misalignment", [])
                novelty_concerns = validation_json.get("novelty_concerns", [])
                improvement_recommendations = validation_json.get("improvement_recommendations", ["Improve experiment design"])
                decision_rationale = validation_json.get("decision_rationale", "")
                
                # MODERATE VALIDATION ENFORCEMENT for literature-grounded experiments
                min_score_met = all(score >= 0.75 for score in detailed_scores.values())  # Increased from 0.70
                if (overall_score < 0.80 or  # Increased from 0.75
                    not min_score_met or
                    len(critical_issues) > 1):  # Reduced from 2 to be stricter
                    validation_result = "FAIL"
               
                _write_stream(f"Validation Result: {validation_result}")
                
                for dimension, score in detailed_scores.items():
                    status = "PASS" if score >= 0.85 else "FAIL"
                    _write_stream(f"   {dimension.replace('_', ' ').title()}: {score:.2f}/1.0 {status}")
                

            except (json.JSONDecodeError, Exception) as e:
                print(f"LLM validation failed: {str(e)}")
                validation_result = "FAIL"
                overall_score = 0.0
                validation_json = {
                    "validation_result": "FAIL",
                    "overall_score": 0.0,
                    "detailed_scores": {"research_direction_alignment": 0.0, "novelty_potential": 0.0, "justification_quality": 0.0},
                    "critical_issues": ["LLM validation parsing failed"],
                    "direction_misalignment": [],
                    "novelty_concerns": [],
                    "improvement_recommendations": ["Fix validation system", "Regenerate experiments"],
                    "decision_rationale": f"Validation failed due to error: {str(e)}"
                }

        # STRICT VALIDATION CRITERIA with LLM results and special handling for no papers
        papers_available = len(validated_papers) > 0

        # DEBUG: Log validation state
        #print(f"🔍 VALIDATION DEBUG: papers_available={papers_available}, validation_result='{validation_result}', overall_score={overall_score:.3f}, current_iteration={current_iteration}")

        # Adjust thresholds based on paper availability - MODERATE STRICTNESS for literature-grounded experiments
        if papers_available:
            min_score_required = 0.80  # Increased from 0.75 for better quality
            pass_condition = validation_result == "PASS" and overall_score >= min_score_required
        else:
            min_score_required = 0.70  # Slightly increased from 0.65
            pass_condition = validation_result == "PASS" and overall_score >= min_score_required
        
        if pass_condition:
            context_note = "with literature context" if papers_available else "without literature context (relaxed criteria)"
            _write_stream(f"Experiment suggestions validation PASSED (LLM score: {overall_score:.2f}/{min_score_required}, {context_note})")
            state["next_node"] = "END"
            state["experiments_validation_decision"] = "PASS"
            
            # Create final outputs with LLM validation results
            final_outputs = {
                "markdown": experiment_suggestions,
                "summary": f"""
# Experiment Suggestion Summary

**Research Direction**: {experiment_summary.get('research_direction', 'Unknown')}
**Total Experiments Designed**: {total_experiments}
**Papers Referenced**: {papers_used}
**LLM Validation Score**: {overall_score:.2f}/1.0
**Validation Status**: {validation_result}
**Generation Status**: Success

## LLM Validation Results
- Overall Score: {overall_score:.2f}/1.0 (Required: ≥0.80)
- Research Direction Alignment: {detailed_scores.get('research_direction_alignment', 0):.2f}/1.0
- Novelty Potential: {detailed_scores.get('novelty_potential', 0):.2f}/1.0
- Justification Quality: {detailed_scores.get('justification_quality', 0):.2f}/1.0

## Quick Overview
{experiment_suggestions[:500]}...

*For complete experiment details, see the full markdown output.*
                """.strip(),
                "json": json.dumps({
                    "experiment_suggestions": experiment_suggestions,
                    "summary": experiment_summary,
                    "llm_validation_results": validation_json,
                    "metadata": {
                        "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                        "iteration": state.get("current_experiment_iteration", 1),
                        "total_papers_found": len(state.get("experiment_papers", [])),
                        "validated_experiment_papers": len(state.get("validated_experiment_papers", []))
                    }
                }, indent=2)
            }
            
            state["final_outputs"] = final_outputs
            
            # Add completion message
            state["messages"].append(
                AIMessage(content=f"Experiment suggestion workflow completed successfully. LLM validation score: {overall_score:.2f}/1.0. Comprehensive experiment designs are ready for implementation.")
            )
        elif current_iteration < 5:  # Allow more retries for experiment generation
            threshold_note = f"minimum {min_score_required:.2f} required" + (" (relaxed - no papers)" if not papers_available else " (strict - with papers)")
            _write_stream(f"Experiment suggestions failed LLM validation - Score: {overall_score:.2f}/1.0 ({threshold_note})")
            state["next_node"] = "suggest_experiments_tree_2"
            state["experiments_validation_decision"] = "RETRY"
            
            # Provide detailed LLM feedback for improvement
            feedback_parts = [f"LLM validation score: {overall_score:.2f}/1.0 (need 0.90+)"]
            if critical_issues:
                feedback_parts.append(f"Critical issues: {'; '.join(critical_issues[:3])}")
            if direction_misalignment:
                feedback_parts.append(f"Direction misalignment: {'; '.join(direction_misalignment[:2])}")
            if novelty_concerns:
                feedback_parts.append(f"Novelty concerns: {'; '.join(novelty_concerns[:2])}")
            if improvement_recommendations:
                feedback_parts.append(f"Recommendations: {'; '.join(improvement_recommendations[:3])}")
            
            current_feedback = " | ".join(feedback_parts)
            state["validation_feedback"] = current_feedback
            
            # ACCUMULATE PAST MISTAKES FOR LEARNING
            past_mistakes = state.get("past_experiment_mistakes", [])
            
            # Create detailed mistake record for this iteration
            mistake_record = {
                "iteration": current_iteration,  # The iteration that just failed validation
                "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                "validation_score": overall_score,
                "critical_issues": critical_issues,
                "direction_misalignment": direction_misalignment,
                "novelty_concerns": novelty_concerns,
                "improvement_recommendations": improvement_recommendations,
                "experiment_summary": experiment_summary,
                "feedback_summary": current_feedback
            }
            
            past_mistakes.append(mistake_record)
            state["past_experiment_mistakes"] = past_mistakes
            
            #print(f" Accumulated {len(past_mistakes)} past mistake records for learning")
            
        else:
            # Force pass after 3 iterations
            print(f"⚠️ Maximum iterations reached. Forcing completion with LLM score: {overall_score:.2f}/1.0")
            state["next_node"] = "END"
            state["experiments_validation_decision"] = "FORCE_PASS"
            
            # Create outputs even for forced pass
            state["final_outputs"] = {
                "markdown": experiment_suggestions,
                "summary": f"Experiment suggestions generated but failed LLM validation (Score: {overall_score:.2f}/1.0) - Forced completion after 3 iterations",
                "llm_validation_results": validation_json,
                "forced_completion": True
            }
        
    except Exception as e:
        error_msg = f"Experiment validation failed: {str(e)}"
        print(f"❌ {error_msg}")
        state["errors"].append(error_msg)
        state["next_node"] = "END"  # Exit on error
        state["experiments_validation_decision"] = "ERROR"
    
    return state


# ==================================================================================
# WORKFLOW CONTROL FUNCTIONS
# ==================================================================================


def _debug_validation_routing(state: ExperimentSuggestionState) -> str:
    """Debug routing function for validate_experiments_tree_2 node."""
    next_node = state.get("next_node", "END")
    validation_decision = state.get("experiments_validation_decision", "PASS")
    current_iteration = state.get("current_experiment_iteration", 0)

    # SAFETY CHECK: Prevent infinite loops - align with validation logic (max 5 iterations)
    MAX_ITERATIONS = 5
    if current_iteration >= MAX_ITERATIONS:
        _write_stream(f"Maximum iterations reached ({MAX_ITERATIONS}), forcing workflow END to prevent infinite recursion")
        return "END"

    if next_node == "END" or validation_decision in ["PASS", "FORCE_PASS"]:
        return "END"
    else:
        return "suggest_experiments_tree_2"
    
    
def _should_continue_with_papers(state: ExperimentSuggestionState) -> str:
    """Determine whether to continue with current papers or search again."""
    validation_decision = state.get("experiment_paper_validation_decision", "continue")
    search_iteration = state.get("experiment_search_iteration", 0)

    # Safety check: After 3 iterations, force continue
    if search_iteration >= 3:
        _write_stream("Maximum search iterations reached (3), forcing continue...")
        return "continue"

    # Map validation decisions to workflow routing
    if validation_decision == "search_backup":
        _write_stream("Searching for additional papers...")
        return "search_backup"
    elif validation_decision == "search_new":
        _write_stream("Starting new paper search...")
        return "search_new"
    else:
        _write_stream("Continuing with current papers...")
        return "continue"


def _should_proceed_with_direction(state: ExperimentSuggestionState) -> str:
    """Determine whether to proceed with the research direction."""
    validation_decision = state.get("direction_validation_decision", "PASS")
    direction_iterations = state.get("direction_iterations", [])

    # Safety check: After 3 iterations, force proceed
    if len(direction_iterations) >= 3:
        _write_stream("Maximum direction validation iterations reached (3), forcing proceed...")
        return "proceed"

    if validation_decision == "PASS":
        return "proceed"
    else:
        return "revise_direction"


def _should_proceed_with_analysis(state: ExperimentSuggestionState) -> str:
    """Determine whether to proceed with analysis or revise it."""
    next_node = state.get("next_node", "decide_research_direction")
    analysis_iterations = state.get("analysis_iterations", [])
    validation_result = state.get("analysis_validation_decision", "FAIL")

    # Safety check: After 3 iterations, force proceed to prevent infinite loops
    if len(analysis_iterations) >= 3:
        _write_stream("Maximum analysis iterations reached (3), forcing proceed to research direction...")
        return "decide_research_direction"
    elif validation_result == "PASS":
        return "decide_research_direction"
    else:
        return "analyze_findings"

    return next_node
# ==================================================================================
# WORKFLOW GRAPH BUILDER
# ==================================================================================
def _build_analyze_and_suggest_experiment_graph() -> StateGraph:
    """Analyze the results and suggest experiments based on findings."""
    workflow = StateGraph(ExperimentSuggestionState)

    # Add nodes for experiment suggestion workflow
    workflow.add_node("analyze_findings", _analyze_experiment_findings_node)
    workflow.add_node("validate_analysis", _validate_analysis_node)
    workflow.add_node("decide_research_direction", _decide_research_direction_node)
    workflow.add_node("validate_research_direction", _validate_research_direction_node)
    workflow.add_node("generate_experiment_search_query", _generate_experiment_search_query_node)
    workflow.add_node("search_experiment_papers", _search_experiment_papers_node)
    workflow.add_node("validate_experiment_papers", _validate_experiment_papers_node)
    workflow.add_node("distill_paper_methodologies", _distill_paper_methodologies_node)
    
    # NEW CLEAN ARCHITECTURE (RECOMMENDED - no dual edges, no state conflicts)
    workflow.add_node("suggest_experiments_tree_2", _suggest_experiments_tree_2_node)
    workflow.add_node("validate_experiments_tree_2", _validate_experiments_tree_2_node)

    # Define the flow
    workflow.set_entry_point("analyze_findings")
    workflow.add_edge("analyze_findings", "validate_analysis")
    
    # Conditional edge after analysis validation - use next_node field
    workflow.add_conditional_edges(
        "validate_analysis",
        _should_proceed_with_analysis,
        {
            "decide_research_direction": "decide_research_direction",  # Analysis is valid, continue
            "analyze_findings": "analyze_findings"  # Analysis needs improvement, iterate
        }
    )
    
    workflow.add_edge("decide_research_direction", "validate_research_direction")
    
    # Conditional edge after research direction validation - use next_node field
    workflow.add_conditional_edges(
        "validate_research_direction",
        lambda state: state.get("next_node", "decide_research_direction"),
        {
            "generate_experiment_search_query": "generate_experiment_search_query",  # Direction is valid, continue
            "decide_research_direction": "decide_research_direction"  # Direction needs refinement, iterate
        }
    )
    
    workflow.add_edge("generate_experiment_search_query", "search_experiment_papers")
    workflow.add_edge("search_experiment_papers", "validate_experiment_papers")
    
    # Conditional edge after validation - use next_node field  
    workflow.add_conditional_edges(
        "validate_experiment_papers",
        lambda state: state.get("next_node", "distill_paper_methodologies"),  # DEFAULT: Route to distillation step
        {
            "distill_paper_methodologies": "distill_paper_methodologies",  # NEW: Distill methodologies first
            "suggest_experiments_tree_2": "distill_paper_methodologies",    # Fallback: also route to distillation
            "search_experiment_papers": "search_experiment_papers", # Keep current papers, search for backup
            "generate_experiment_search_query": "generate_experiment_search_query"  # Start fresh with new search query
        }
    )
    
    # Add edge from distillation to experiment suggestion
    workflow.add_edge("distill_paper_methodologies", "suggest_experiments_tree_2")
    workflow.add_conditional_edges(
        "suggest_experiments_tree_2",
        lambda state: state.get("next_node", "validate_experiments_tree_2"),  # Default to validation
        {
            "validate_experiments_tree_2": "validate_experiments_tree_2",
            "END": END
        }
    )
    
    # NEW CLEAN CONDITIONAL EDGE (ACTIVE - no state conflicts)
    workflow.add_conditional_edges(
        "validate_experiments_tree_2",
        lambda state: _debug_validation_routing(state),
        {
            "END": END,  # Experiments are valid, finish workflow
            "suggest_experiments_tree_2": "suggest_experiments_tree_2"  # Loop back with feedback
        }
    )

    return workflow.compile()
# ==================================================================================
# MAIN WORKFLOW RUNNER
# ==================================================================================

async def run_experiment_suggestion_workflow_nonstream(
    user_prompt: str,
    experimental_results: Dict[str, Any] = None,
    uploaded_data: List[str] = None,
    file_path: str = None
) -> Dict[str, Any]:
    """
    Compile and run the complete experiment suggestion workflow.
    
    Args:
        user_prompt: The user's research query/context
        experimental_results: Optional dictionary containing experimental data and results
        uploaded_data: Optional list of uploaded file contents
        file_path: Optional path to a file to read and include as input data
        
    Returns:
        Dictionary containing the final workflow state with results
    """
    # Move all imports and initialization inside the function
    try:
        import asyncio
        import openai
        from utils.arxiv_paper_utils import ArxivPaperProcessor
    except ImportError as e:
        error_msg = f"Failed to import required modules: {str(e)}. Please ensure all dependencies are installed."
        print(f"❌ {error_msg}")
        return {
            "workflow_successful": False,
            "error": error_msg,
            "error_type": "ImportError",
            "original_prompt": user_prompt
        }
    
    try:
        # Load configuration
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise ValueError("OPENAI_API_KEY not found in environment variables. Please ensure it is set.")

        base_url = os.getenv("BASE_URL")
        model = os.getenv("DEFAULT_MODEL") or "gemini/gemini-2.5-flash"
        model_cheap = "gemini/gemini-2.5-flash-lite"
        model_expensive = "gemini/gemini-2.5-pro"

        # Initialize dependencies
        try:
            client = openai.OpenAI(api_key=api_key, base_url=base_url)
        except Exception as e:
            raise ValueError(f"Failed to initialize OpenAI client: {str(e)}. Please check your API key and base URL configuration.")
        
        try:
            arxiv_processor = ArxivPaperProcessor(llm_client=client, model_name=model_cheap)
        except Exception as e:
            raise ValueError(f"Failed to initialize ArxivPaperProcessor: {str(e)}. Please check the ArxivPaperProcessor implementation.")
        
    except ValueError as e:
        print(f"❌ Configuration Error: {str(e)}")
        return {
            "workflow_successful": False,
            "error": str(e),
            "error_type": "ConfigurationError",
            "original_prompt": user_prompt
        }
    except Exception as e:
        error_msg = f"Unexpected error during initialization: {str(e)}"
        print(f"❌ {error_msg}")
        return {
            "workflow_successful": False,
            "error": error_msg,
            "error_type": type(e).__name__,
            "original_prompt": user_prompt
        }
    
    print("🧪 Starting Experiment Suggestion Workflow...")
    print(f"📝 User Prompt: {user_prompt}")
    print(f"🔬 Experimental Results: {len(experimental_results) if experimental_results else 0} data points")
    print(f"🤖 Model: {model}")
    if file_path:
        print(f"📁 File Input: {file_path}")
    print("=" * 80)

    # Handle file input if provided
    file_content: List[str] = []
    file_warnings: List[str] = []
    if file_path:
        try:
            print(f"📖 Reading file: {file_path}")
            file_content, file_warnings = _load_text_file_safely(file_path)
            if file_content:
                print(f"✅ File loaded successfully ({len(file_content[0])} characters)")
        except FileNotFoundError:
            print(f"❌ Error: File not found - {file_path}")
            raise
        except Exception as e:
            print(f"❌ Error reading file: {str(e)}")
            raise

        for warning in file_warnings:
            print(warning)

    # Combine uploaded_data with file_content if both are provided
    combined_uploaded_data = (uploaded_data or []) + file_content
    
    try:
        # Build the workflow graph
        workflow_graph = _build_analyze_and_suggest_experiment_graph()
        print("✅ Workflow graph compiled successfully")
        
        # Initialize the state with all required fields
        initial_state = {
            # Core workflow data
            "messages": [],
            "original_prompt": user_prompt,
            "uploaded_data": combined_uploaded_data,
            "current_step": "starting",
            "errors": [],
            "workflow_type": "experiment_suggestion",
            
            # Dependencies
            "client": client,
            "model": model,
            "arxiv_processor": arxiv_processor,
            
            # Input data
            "experimental_results": experimental_results or {},
            "findings_analysis": {},
            "research_context": {},
            
            # Processing state
            "analysis_completed": False,
            "experiment_categories": [],
            "experiment_papers": [],
            "experiment_search_query": "",
            "experiment_search_iteration": 0,
            "experiment_validation_results": {},
            "experiment_paper_validation_decision": "",
            "experiment_validation_decision": "",
            
            # New validation fields for the exact workflow structure
            "analysis_validation_decision": "",
            "direction_validation_decision": "",
            "paper_validation_decision": "",
            "experiments_validation_decision": "",
            "validation_feedback": "",
            
            "experiment_iterations": [],
            "research_direction": {},
            "validated_experiment_papers": [],
            "validated_experiment_papers": [],  # Add this key that the clean architecture uses
            "distilled_methodologies": {},       # Distilled methodology content from papers
            "current_experiment_iteration": 0,
            "iteration_from_state": 0,
            "analysis_iterations": [],  # Track analysis validation iterations (list for history)
            "direction_iterations": [],  # Track research direction validation iterations (list for history)
            
            # Issue tracking
            "past_fixed_issues": [],
            "past_unresolved_issues": [],
            "most_recent_generation_issues": [],
            "cumulative_validation_feedback": [],
            
            # Solved issues tracking
            "solved_issues_history": [],
            "current_solved_issues": [],
            "validation_issue_patterns": {},
            "generation_feedback_context": "",
            
            # Output
            "experiment_suggestions": "",
            "experiment_summary": {},
            "next_node": "",
            "literature_context": "",
            "suggestion_source": "",
            "prioritized_experiments": [],
            "implementation_roadmap": {},
            "final_outputs": {}
        }
        
        print("🔄 Running workflow...")
        
        # Run the workflow with increased recursion limit to prevent infinite loops
        final_state = await workflow_graph.ainvoke(initial_state, config={"recursion_limit": 50})
        
        print("\n" + "=" * 80)
        print("🎉 WORKFLOW COMPLETED SUCCESSFULLY!")
        print("=" * 80)
        
        # Extract and display key results
        if final_state.get("experiment_suggestions"):
            print("✅ Experiment suggestions generated successfully")
            suggestions = final_state.get("experiment_suggestions", "")
            if suggestions:
                print(f"\n📋 EXPERIMENT SUGGESTIONS PREVIEW:")
                print("-" * 40)
                print(suggestions[:500] + "..." if len(suggestions) > 500 else suggestions)
        else:
            print("⚠️ Experiment suggestions may be incomplete")
        
        # Display any errors
        if final_state.get("errors"):
            print(f"\n⚠️ Errors encountered: {len(final_state['errors'])}")
            for i, error in enumerate(final_state["errors"][-3:], 1):  # Show last 3 errors
                print(f"  {i}. {error}")
        
        # Display workflow statistics
        print(f"\n📊 WORKFLOW STATISTICS:")
        print(f"   - Papers found: {len(final_state.get('experiment_papers', []))}")
        print(f"   - Papers validated: {len(final_state.get('validated_papers', []))}")
        print(f"   - Search iterations: {final_state.get('experiment_search_iteration', 0)}")
        print(f"   - Experiment iterations: {final_state.get('current_experiment_iteration', 0)}")
        
        return final_state
        
    except Exception as e:
        print(f"\n❌ WORKFLOW FAILED: {str(e)}")
        print("Full error traceback:")
        traceback.print_exc()
        
        # Return error state
        return {
            "workflow_successful": False,
            "error": str(e),
            "error_type": type(e).__name__,
            "original_prompt": user_prompt,
            "experimental_results": experimental_results
        }





async def run_experiment_suggestion_workflow(
    user_prompt: str,
    experimental_results: Dict[str, Any] = None,
    uploaded_data: List[str] = None,
    file_path: str = None,
    streaming: bool = False
) -> Dict[str, Any]:
    """
    Compile and run the complete experiment suggestion workflow.
    
    Args:
        user_prompt: The user's research query/context
        experimental_results: Optional dictionary containing experimental data and results
        uploaded_data: Optional list of uploaded file contents
        file_path: Optional path to a file to read and include as input data
        
    Returns:
        Dictionary containing the final workflow state with results
    """
    # Move all imports and initialization inside the function
    try:
        import asyncio
        import openai
        from utils.arxiv_paper_utils import ArxivPaperProcessor
    except ImportError as e:
        error_msg = f"Failed to import required modules: {str(e)}. Please ensure all dependencies are installed."
        print(f"❌ {error_msg}")
        return {
            "workflow_successful": False,
            "error": error_msg,
            "error_type": "ImportError",
            "original_prompt": user_prompt
        }
    
    try:
        # Load configuration
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise ValueError("OPENAI_API_KEY not found in environment variables. Please ensure it is set.")

        base_url = os.getenv("BASE_URL")
        model = os.getenv("DEFAULT_MODEL") or "gemini/gemini-2.5-flash"
        model_cheap = "gemini/gemini-2.5-flash-lite"
        model_expensive = "gemini/gemini-2.5-pro"

        # Initialize dependencies
        try:
            client = openai.OpenAI(api_key=api_key, base_url=base_url)
        except Exception as e:
            raise ValueError(f"Failed to initialize OpenAI client: {str(e)}. Please check your API key and base URL configuration.")
        
        try:
            arxiv_processor = ArxivPaperProcessor(llm_client=client, model_name=model_cheap)
        except Exception as e:
            raise ValueError(f"Failed to initialize ArxivPaperProcessor: {str(e)}. Please check the ArxivPaperProcessor implementation.")
        
    except ValueError as e:
        print(f"❌ Configuration Error: {str(e)}")
        return {
            "workflow_successful": False,
            "error": str(e),
            "error_type": "ConfigurationError",
            "original_prompt": user_prompt
        }
    except Exception as e:
        error_msg = f"Unexpected error during initialization: {str(e)}"
        print(f"❌ {error_msg}")
        return {
            "workflow_successful": False,
            "error": error_msg,
            "error_type": type(e).__name__,
            "original_prompt": user_prompt
        }
    
    print("🧪 Starting Experiment Suggestion Workflow...")
    print(f"📝 User Prompt: {user_prompt}")
    print(f"🔬 Experimental Results: {len(experimental_results) if experimental_results else 0} data points")
    print(f"🤖 Model: {model}")
    if file_path:
        print(f"📁 File Input: {file_path}")
    print("=" * 80)

    # Handle file input if provided
    file_content: List[str] = []
    file_warnings: List[str] = []
    if file_path:
        try:
            print(f" Reading file: {file_path}")
            file_content, file_warnings = _load_text_file_safely(file_path)
            if file_content:
                print(f"File loaded successfully ({len(file_content[0])} characters)")
        except FileNotFoundError:
            print(f"Error: File not found - {file_path}")
            raise
        except Exception as e:
            print(f"Error reading file: {str(e)}")
            raise

        for warning in file_warnings:
            print(warning)

    # Combine uploaded_data with file_content if both are provided
    combined_uploaded_data = (uploaded_data or []) + file_content
    
    try:
        # Build the workflow graph
        workflow_graph = _build_analyze_and_suggest_experiment_graph()
        print("✅ Workflow graph compiled successfully")
        
        # Initialize the state with all required fields
        initial_state = {
            # Core workflow data
            "messages": [],
            "original_prompt": user_prompt,
            "uploaded_data": combined_uploaded_data,
            "current_step": "starting",
            "errors": [],
            "workflow_type": "experiment_suggestion",
            
            # Dependencies
            "client": client,
            "model": model,
            "arxiv_processor": arxiv_processor,
            
            # Input data
            "experimental_results": experimental_results or {},
            "findings_analysis": {},
            "research_context": {},
            
            # Processing state
            "analysis_completed": False,
            "experiment_categories": [],
            "experiment_papers": [],
            "experiment_search_query": "",
            "experiment_search_iteration": 0,
            "experiment_validation_results": {},
            "experiment_paper_validation_decision": "",
            "experiment_validation_decision": "",
            
            # New validation fields for the exact workflow structure
            "analysis_validation_decision": "",
            "direction_validation_decision": "",
            "paper_validation_decision": "",
            "experiments_validation_decision": "",
            "validation_feedback": "",
            
            "experiment_iterations": [],
            "research_direction": {},
            "validated_experiment_papers": [],
            "validated_experiment_papers": [],  # Add this key that the clean architecture uses
            "distilled_methodologies": {},       # Distilled methodology content from papers
            "current_experiment_iteration": 0,
            "iteration_from_state": 0,
            "analysis_iterations": [],  # Track analysis validation iterations (list for history)
            "direction_iterations": [],  # Track research direction validation iterations (list for history)
            
            # Issue tracking
            "past_fixed_issues": [],
            "past_unresolved_issues": [],
            "most_recent_generation_issues": [],
            "cumulative_validation_feedback": [],
            
            # Solved issues tracking
            "solved_issues_history": [],
            "current_solved_issues": [],
            "validation_issue_patterns": {},
            "generation_feedback_context": "",
            
            # Output
            "experiment_suggestions": "",
            "experiment_summary": {},
            "next_node": "",
            "literature_context": "",
            "suggestion_source": "",
            "prioritized_experiments": [],
            "implementation_roadmap": {},
            "final_outputs": {}
        }
        
        print("🔄 Running workflow...")
        
        if not streaming:
            final_state = await workflow_graph.ainvoke(initial_state)
            return final_state.get("validate_experiments_tree_2", final_state)
        # Run the workflow with increased recursion limit to prevent infinite loops
        async def _stream():
            final_data = None  # track last update

            async for chunk in workflow_graph.astream(initial_state, stream_mode=["updates","custom"]):
                stream_mode, data = chunk

                # Debugging / logging (optional, can remove to stay "silent")
                if stream_mode == "updates":
                    key = list(data.keys())[0] if data else None
                    print(f"Node Complete: {key}.")
                    print("-" * 20)
                elif stream_mode == "custom" and data.get("status"):
                    print(f"Updates: {data['status']}")

                # Stream intermediate updates
                yield data
                final_data = data

            # After loop ends, yield final state only if it has "validate_experiments_tree_2"
            if final_data and "validate_experiments_tree_2" in final_data:
                yield final_data["validate_experiments_tree_2"]

        return _stream()

    except Exception as e:
        print(f"\n❌ WORKFLOW FAILED: {str(e)}")
        print("Full error traceback:")
        traceback.print_exc()
        
        # Return error state
        return {
            "workflow_successful": False,
            "error": str(e),
            "error_type": type(e).__name__,
            "original_prompt": user_prompt,
            "experimental_results": experimental_results
        }

async def test_experiment_stream():
    #import asyncio
    from dotenv import load_dotenv
    load_dotenv()  # Load environment variables from .env file if present
    test_prompt = "I have completed initial experiments on image classification with CNNs. Need suggestions for follow-up experiments to improve model performance and generalization."
    data_dir=r"C:\Users\Jacobs laptop\Downloads\AETHER Hackathon.xlsx"
    final_result = None
    async for result in await run_experiment_suggestion_workflow(test_prompt,file_path=data_dir, streaming=True):
        final_result = result  # keep overwriting, so last one wins
        
    result = final_result
    
    # FOR NON STREAMING
    #result = await run_model_suggestion_workflow("Find models for X-rays", streaming=False)
    
   
    return result
    
    
if __name__ == "__main__":
    """
    Main entry point for testing the workflow.
    """
    # Run test
    import asyncio
    from dotenv import load_dotenv
    load_dotenv()  # Load environment variables from .env file if present
    result = asyncio.run(test_experiment_stream())
    print("Final result:", "Success" if result.get("experiment_suggestions") else "Failed")


